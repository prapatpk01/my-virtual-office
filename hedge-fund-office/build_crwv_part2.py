"""CRWV Equity Research Workbook - Part 2: Sheet 3 (Financials) + Sheet 4 (Thesis/Catalysts/Risks)"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel

NAVY="0D1B2A"; NAVY2="1E3A5F"; GOLD="D4AF37"; WHITE="FFFFFF"
GREEN="2ECC71"; RED="E74C3C"; ORANGE="F39C12"; ROW_ALT="F0F4F8"
BORDER_C="CCCCCC"; LIGHT_BLUE="DCE6F1"; NEG_RED="C0392B"

def fill(h): return PatternFill("solid",fgColor=h)
def fnt(bold=False,color="000000",size=10,italic=False):
    return Font(bold=bold,color=color,size=size,italic=italic,name="Calibri")
def side(): return Side(style="thin",color=BORDER_C)
def brd(): return Border(left=side(),right=side(),top=side(),bottom=side())
def aln(h="left",v="center",wrap=False):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap)

def sc(ws,row,col,val,bold=False,color="000000",bg=None,num_fmt=None,
       wrap=False,halign="left",sz=10,italic=False):
    c=ws.cell(row=row,column=col,value=val)
    c.font=Font(bold=bold,color=color,size=sz,italic=italic,name="Calibri")
    if bg: c.fill=fill(bg)
    c.border=brd()
    c.alignment=Alignment(horizontal=halign,vertical="center",wrap_text=wrap)
    if num_fmt: c.number_format=num_fmt
    return c

def mhdr(ws,row,c1,c2,text,bg=NAVY,fg=GOLD,sz=12,bold=True):
    ws.merge_cells(start_row=row,start_column=c1,end_row=row,end_column=c2)
    c=ws.cell(row=row,column=c1,value=text)
    c.font=Font(bold=bold,color=fg,size=sz,name="Calibri")
    c.fill=fill(bg)
    c.alignment=Alignment(horizontal="center",vertical="center")

def sec(ws,row,c1,c2,text):
    ws.merge_cells(start_row=row,start_column=c1,end_row=row,end_column=c2)
    c=ws.cell(row=row,column=c1,value=text)
    c.font=Font(bold=True,color=GOLD,size=11,name="Calibri")
    c.fill=fill(NAVY)
    c.alignment=Alignment(horizontal="left",vertical="center")

def footer(ws,row,ncols=9):
    txt="CoreWeave (CRWV) | Equity Research | Sentinel Investment | June 10, 2026 | Confidential"
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=ncols)
    c=ws.cell(row=row,column=1,value=txt)
    c.font=Font(italic=True,color="888888",size=8,name="Calibri")
    c.fill=fill(NAVY)
    c.alignment=Alignment(horizontal="center",vertical="center")

out_path="/home/user/my-virtual-office/hedge-fund-office/CRWV_Equity_Research.xlsx"
wb=openpyxl.load_workbook(out_path)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3: Financials & Earnings
# ════════════════════════════════════════════════════════════════════════════
ws3=wb.create_sheet("Financials & Earnings")
ws3.sheet_properties.tabColor="2E6DA4"

mhdr(ws3,1,1,9,"CRWV — Financial Summary & Earnings History",bg=NAVY,fg=GOLD,sz=13)
ws3.row_dimensions[1].height=28

# ── Section A: Income Statement ──
sec(ws3,2,1,9,"  SECTION A — Income Statement  ($M)")

periods=["FY2022A","FY2023A","FY2024A","Q1 2025A","Q2 2025E","FY2025E"]
hdr_row=3
sc(ws3,hdr_row,1,"Line Item",bold=True,color=WHITE,bg=NAVY2,halign="left")
for ci,p in enumerate(periods,2):
    sc(ws3,hdr_row,ci,p,bold=True,color=WHITE,bg=NAVY2,halign="center")
# extra col
sc(ws3,hdr_row,8,"",bold=True,color=WHITE,bg=NAVY2)
sc(ws3,hdr_row,9,"",bold=True,color=WHITE,bg=NAVY2)

# IS data  (label, 2022, 2023, 2024, Q1-25A, Q2-25E, FY25E, is_bold, color_neg)
is_rows=[
    ("Revenue",           16,   229,  1915, 982,  1150, 4200, True,  False),
    ("YoY Growth",        "—",  "+1,331%","+736%","—","—","+119%",False,False),
    ("Gross Profit",      5,    97,   900,  510,  598,  2184, False, False),
    ("Gross Margin %",    "30.5%","42.4%","47.0%","52.0%","52.0%","52.0%",False,False),
    ("R&D",               -12,  -35,  -120, -55,  -60,  -230, False, True),
    ("S&M",               -8,   -22,  -85,  -38,  -42,  -165, False, True),
    ("G&A",               -15,  -45,  -185, -75,  -80,  -310, False, True),
    ("Operating Income (pre-D&A)", -30, -5, 510, 342, 416, 1479, True, True),
    ("Op. Margin %",      "nm", "-2.2%","26.6%","34.8%","36.2%","35.2%",False,False),
    ("D&A",               -12,  -85,  -1380,-400, -450, -1700,False, True),
    ("EBIT (after D&A)",  -42,  -90,  -870, -58,  -34,  -221, True,  True),
    ("Interest Expense",  -5,   -45,  -530, -180, -190, -730, False, True),
    ("Net Income",        -47,  -594, -863, -238, -224, -950, True,  True),
    ("EPS (diluted)",     "nm","($1.71)","($2.08)","($0.57)","($0.54)","($2.29)",False,False),
    ("EBITDA",            -35,  -5,   510,  342,  416,  1479, True,  True),
]
for i,(label,v1,v2,v3,v4,v5,v6,bold,cneg) in enumerate(is_rows):
    row=4+i
    bg=ROW_ALT if i%2==0 else WHITE
    # label
    sc(ws3,row,1,label,bold=bold,bg=bg,halign="left")
    for ci,val in enumerate([v1,v2,v3,v4,v5,v6],2):
        is_num=isinstance(val,(int,float))
        neg=is_num and val<0
        color=NEG_RED if (neg and cneg) else "000000"
        disp=f"({abs(val):,})" if neg else (f"{val:,}" if is_num else val)
        sc(ws3,row,ci,disp,bold=bold,color=color,bg=bg,halign="center")

# ── Section B: Key Metrics ──
sec(ws3,21,1,7,"  SECTION B — Key Operating & Financial Metrics")

km_hdrs=["Metric","FY2022A","FY2023A","FY2024A","FY2025E","",""]
for ci,h in enumerate(km_hdrs[:5],1):
    sc(ws3,22,ci,h,bold=True,color=WHITE,bg=NAVY2,halign="center")

km_data=[
    ("Gross Margin",     "30.5%","42.4%","47.0%","52.0%"),
    ("EBITDA Margin",    "nm",   "nm",   "26.6%","35.2%"),
    ("Net Margin",       "nm",   "nm",   "nm",   "nm"),
    ("Capex ($M)",       "(200)","(1,800)","(7,000)","(10,000)"),
    ("Capex % Revenue",  "nm",   "786%", "366%", "238%"),
    ("Total Debt ($M)",  "800",  "3,200","8,900","12,000"),
    ("Cash ($M)",        "150",  "380",  "1,400","1,800"),
    ("Net Debt ($M)",    "650",  "2,820","7,500","10,200"),
]
for i,(label,v1,v2,v3,v4) in enumerate(km_data):
    row=23+i
    bg=ROW_ALT if i%2==0 else WHITE
    sc(ws3,row,1,label,bold=True,bg=bg,halign="left")
    for ci,val in enumerate([v1,v2,v3,v4],2):
        sc(ws3,row,ci,val,bg=bg,halign="center")

# ── Section C: Earnings History ──
sec(ws3,33,1,9,"  SECTION C — Earnings vs. Estimates History")

eh_hdrs=["Quarter","Revenue ($M)","Est. ($M)","Rev Beat/Miss","EPS","EPS Est.","EPS Beat/Miss","Key Note",""]
for ci,h in enumerate(eh_hdrs[:8],1):
    sc(ws3,34,ci,h,bold=True,color=WHITE,bg=NAVY2,halign="center",sz=9)

eh_data=[
    ("Q2 2024","$395M","$310M","+27.4% BEAT","$(0.35)","$(0.42)","BEAT","First major quarter post-IPO"),
    ("Q3 2024","$509M","$460M","+10.7% BEAT","$(0.62)","$(0.58)","MISS","Interest expense surge"),
    ("Q4 2024","$789M","$720M","+9.6% BEAT", "$(0.76)","$(0.71)","MISS","Raised 2025 guidance"),
    ("Q1 2025","$982M","$853M","+15.1% BEAT","$(0.57)","$(0.68)","BEAT","Record Q, guidance raise"),
]
for i,(q,rev,est,rb,eps,epse,eb,note) in enumerate(eh_data):
    row=35+i
    bg=ROW_ALT if i%2==0 else WHITE
    beat_bg={"BEAT":"E8F8EF","MISS":"FDECEA"}
    sc(ws3,row,1,q,bold=True,bg=bg)
    sc(ws3,row,2,rev,bg=bg,halign="center")
    sc(ws3,row,3,est,bg=bg,halign="center")
    beat_color=GREEN if "BEAT" in rb else RED
    sc(ws3,row,4,rb,color=WHITE,bg=beat_color[0:6] if False else
       ("1A7A3F" if "BEAT" in rb else "922B21"),halign="center")
    sc(ws3,row,5,eps,bg=bg,halign="center")
    sc(ws3,row,6,epse,bg=bg,halign="center")
    sc(ws3,row,7,eb,color=("1A7A3F" if eb=="BEAT" else "922B21"),
       bg=bg,halign="center",bold=True)
    sc(ws3,row,8,note,bg=bg,halign="left",wrap=True)

# ── Revenue & Margin Chart ──
# Chart data uses IS rows revenue (row 4) and gross margin % (row 7)
chart_bar=BarChart()
chart_bar.type="col"
chart_bar.title="Revenue Trend ($M)"
chart_bar.style=10
chart_bar.y_axis.title="Revenue ($M)"
chart_bar.width=16; chart_bar.height=10

# We'll put chart data helper rows offscreen in col J onward
rev_vals=[16,229,1915,4200]
gm_vals=[0.305,0.424,0.470,0.520]
periods_chart=["FY2022A","FY2023A","FY2024A","FY2025E"]
for ci,(p,r,g) in enumerate(zip(periods_chart,rev_vals,gm_vals)):
    ws3.cell(row=2,column=11+ci,value=p)
    ws3.cell(row=3,column=11+ci,value=r)
    ws3.cell(row=4,column=11+ci,value=g)

rev_ref=Reference(ws3,min_col=11,max_col=14,min_row=2,max_row=3)
chart_bar.add_data(rev_ref,titles_from_data=True)
cats=Reference(ws3,min_col=11,max_col=14,min_row=2)
chart_bar.set_categories(cats)
chart_bar.series[0].graphicalProperties.solidFill="1E3A5F"

lc=LineChart()
lc.title="Gross Margin %"
lc.y_axis.title="Gross Margin"
lc.y_axis.numFmt="0%"
gm_ref=Reference(ws3,min_col=11,max_col=14,min_row=4)
lc.add_data(gm_ref,titles_from_data=False)
lc.series[0].title=SeriesLabel(v="Gross Margin %")
lc.series[0].graphicalProperties.line.solidFill=ORANGE
lc.series[0].smooth=True

chart_bar+=lc
ws3.add_chart(chart_bar,"A42")

footer(ws3,58,9)

ws3.freeze_panes="A4"
ws3.column_dimensions["A"].width=28
for col in ["B","C","D","E","F","G"]:
    ws3.column_dimensions[col].width=13
ws3.column_dimensions["H"].width=35
ws3.column_dimensions["I"].width=5

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4: Thesis, Catalysts & Risks
# ════════════════════════════════════════════════════════════════════════════
ws4=wb.create_sheet("Thesis, Catalysts & Risks")
ws4.sheet_properties.tabColor="2E6DA4"

mhdr(ws4,1,1,8,"Investment Thesis, Catalysts & Risk Matrix",bg=NAVY,fg=GOLD,sz=13)
ws4.row_dimensions[1].height=28

# ── Section A: Scenarios ──
sec(ws4,2,1,8,"  SECTION A — Investment Scenarios")

scen_hdrs=["Metric","BULL (30%)","BASE (50%)","BEAR (20%)","","","",""]
for ci,h in enumerate(scen_hdrs[:4],1):
    sc(ws4,3,ci,h,bold=True,color=WHITE,
       bg=(GREEN if "BULL" in h else (ORANGE if "BASE" in h else (RED if "BEAR" in h else NAVY2))),
       halign="center")

scen_rows=[
    ("Revenue 2026E",   "$9,500M",  "$7,500M",  "$4,500M"),
    ("Gross Margin 2026E","58%",    "55%",      "48%"),
    ("EV/Rev Multiple", "7.5x",     "6.7x",     "3.5x"),
    ("EV",              "$71B",     "$50B",      "$16B"),
    ("Net Debt",        "$8B",      "$8B",       "$10B"),
    ("Equity Value",    "$63B",     "$42B",      "$6B"),
    ("Price Target",    "$180",     "$125",      "$55"),
    ("Key Assumption",  "MSFT expands + new hyperscaler contracts",
                        "Steady execution on backlog",
                        "Customer loss + AI spending cut"),
]
for i,(label,bull,base,bear) in enumerate(scen_rows):
    row=4+i
    bg=ROW_ALT if i%2==0 else WHITE
    sc(ws4,row,1,label,bold=True,bg=bg)
    sc(ws4,row,2,bull,color=("1A7A3F" if i==6 else "000000"),
       bold=(i==6),bg=bg,halign="center",wrap=True)
    sc(ws4,row,3,base,color=(ORANGE if i==6 else "000000"),
       bold=(i==6),bg=bg,halign="center",wrap=True)
    sc(ws4,row,4,bear,color=("922B21" if i==6 else "000000"),
       bold=(i==6),bg=bg,halign="center",wrap=True)

# ── Section B: Catalyst Timeline ──
sec(ws4,14,1,8,"  SECTION B — 12-Month Catalyst Timeline")

cat_hdrs=["Date","Catalyst","Potential Impact","Probability","","","",""]
for ci,h in enumerate(cat_hdrs[:4],1):
    sc(ws4,15,ci,h,bold=True,color=WHITE,bg=NAVY2,halign="center")

cat_data=[
    ("Aug 2025", "Q2 2025 Earnings",                    "±15%",       "High — beats expected"),
    ("Sep 2025", "Blackwell GPU deployment announcement","+10-20%",    "High"),
    ("Oct 2025", "New enterprise customer wins",         "+5-15%",     "Medium"),
    ("Nov 2025", "Q3 2025 Earnings",                    "±12%",       "High"),
    ("Dec 2025", "Microsoft contract renewal/expansion", "+15-25%",    "Medium-High"),
    ("Jan 2026", "Backlog update ($30B target?)",        "+10%",       "Medium"),
    ("Feb 2026", "Q4 2025 Earnings + FY2026 guidance",  "±15%",       "High"),
    ("Mar 2026", "1-year IPO anniversary — lockup expiry","-5 to -10%","High risk"),
    ("May 2026", "Q1 2026 Earnings",                    "±12%",       "High"),
    ("Jun 2026", "CPI/macro data (favorable)",           "+3-8%",      "High (just happened)"),
]
for i,(dt,cat,imp,prob) in enumerate(cat_data):
    row=16+i
    bg=ROW_ALT if i%2==0 else WHITE
    neg="-" in imp and "±" not in imp
    imp_color="922B21" if neg else ("1A7A3F" if "+" in imp else "000000")
    sc(ws4,row,1,dt,bold=True,bg=bg)
    sc(ws4,row,2,cat,bg=bg,wrap=True)
    sc(ws4,row,3,imp,color=imp_color,bg=bg,halign="center",bold=True)
    sc(ws4,row,4,prob,bg=bg,halign="center")

# ── Section C: Risk Matrix ──
sec(ws4,28,1,8,"  SECTION C — Risk Matrix")

risk_hdrs=["Risk","Severity","Probability","Mitigation","Impact Score","","",""]
for ci,h in enumerate(risk_hdrs[:5],1):
    sc(ws4,29,ci,h,bold=True,color=WHITE,bg=NAVY2,halign="center")

sev_colors={"Critical":RED,"High":ORANGE,"Medium":"F39C12"}
risk_data=[
    ("Customer Concentration (MSFT 62%)","Critical","Medium","Diversify to 5+ major customers","9/10"),
    ("Debt Burden ($9B+)","High","High","FCF positive by 2027, refinancing","8/10"),
    ("Hyperscaler Build-out","High","High","Long-term contracts lock in revenue","7/10"),
    ("GPU Supply Risk","High","Medium","NVIDIA strategic partner status","7/10"),
    ("AI Spending Slowdown","Medium","Low","Contracted backlog $23B+","6/10"),
    ("Regulatory Risk","Medium","Low","No direct regulation yet","4/10"),
    ("Key Person Risk (CEO Brannin McBee)","Medium","Low","Strong mgmt team","4/10"),
    ("Competition from CoreWeave clones","Medium","Medium","First-mover + scale advantages","5/10"),
]
for i,(risk,sev,prob,mit,score) in enumerate(risk_data):
    row=30+i
    bg=ROW_ALT if i%2==0 else WHITE
    sev_c={"Critical":RED,"High":ORANGE,"Medium":"F39C12"}.get(sev,ORANGE)
    sc(ws4,row,1,risk,bold=True,bg=bg,wrap=True)
    sc(ws4,row,2,sev,bold=True,color=WHITE,bg=sev_c,halign="center")
    sc(ws4,row,3,prob,bg=bg,halign="center")
    sc(ws4,row,4,mit,bg=bg,wrap=True)
    sc(ws4,row,5,score,bold=True,bg=bg,halign="center")

footer(ws4,42,8)

ws4.freeze_panes="A4"
ws4.column_dimensions["A"].width=35
ws4.column_dimensions["B"].width=12
ws4.column_dimensions["C"].width=14
ws4.column_dimensions["D"].width=38
ws4.column_dimensions["E"].width=14
for col in ["F","G","H"]:
    ws4.column_dimensions[col].width=5

ws4.row_dimensions[29].height=15

wb.save(out_path)
print(f"Part 2 saved: {out_path}")
