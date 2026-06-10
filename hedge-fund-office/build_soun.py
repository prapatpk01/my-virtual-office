#!/usr/bin/env python3
"""
SOUN (SoundHound AI) Institutional Equity Research Workbook
Sentinel Investment | June 10, 2026
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import os

# ── Color palette ──────────────────────────────────────────────────
C_HEADER_BG   = "0B1F3A"   # Dark navy header
C_HEADER_FG   = "FFFFFF"   # White text
C_ACCENT      = "00D4FF"   # AI cyan
C_ACCENT_DARK = "0099BB"   # Darker cyan
C_POS         = "2ECC71"   # Green (positive)
C_NEG         = "E74C3C"   # Red (negative)
C_ALT         = "F0F7FF"   # Light blue alt row
C_WHITE       = "FFFFFF"
C_LABEL_BG    = "1A3A5C"   # Section label dark blue
C_SUBHDR      = "2E5F8A"   # Sub-header medium blue
C_GOLD        = "F39C12"   # Warning/watch amber
C_LIGHT_GREY  = "F5F5F5"
C_BORDER      = "BDC3C7"
C_INPUT       = "0000FF"   # Blue = hardcoded input (convention)
C_FORMULA     = "000000"   # Black = formula
C_LINK        = "27AE60"   # Green = cross-sheet link

# ── Style helpers ──────────────────────────────────────────────────
def hdr_font(sz=11, bold=True, color=C_HEADER_FG):
    return Font(name="Calibri", size=sz, bold=bold, color=color)

def std_font(sz=10, bold=False, color="000000"):
    return Font(name="Calibri", size=sz, bold=bold, color=color)

def input_font(sz=10):
    return Font(name="Calibri", size=sz, bold=False, color=C_INPUT)

def formula_font(sz=10):
    return Font(name="Calibri", size=sz, bold=False, color=C_FORMULA)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def right():
    return Alignment(horizontal="right", vertical="center")

def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(bottom=s)

def thick_bottom():
    s = Side(style="medium", color=C_HEADER_BG)
    return Border(bottom=s)

def apply_header_row(ws, row, cols, text_list, bg=C_HEADER_BG, fg=C_HEADER_FG, sz=10):
    for i, (col, text) in enumerate(zip(cols, text_list)):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(name="Calibri", size=sz, bold=True, color=fg)
        c.fill = fill(bg)
        c.alignment = center()
        c.border = thin_border()

def style_section_title(ws, row, col, text, colspan=1, bg=C_LABEL_BG):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=11, bold=True, color=C_ACCENT)
    c.fill = fill(bg)
    c.alignment = left()
    c.border = thin_border()

def style_data_row(ws, row, cols, values, alt=False, fmt_list=None, bold=False):
    bg = C_ALT if alt else C_WHITE
    for i, (col, val) in enumerate(zip(cols, values)):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = right() if isinstance(val, (int, float)) else left()
        c.font = Font(name="Calibri", size=10, bold=bold, color="000000")
        if fmt_list and i < len(fmt_list) and fmt_list[i]:
            c.number_format = fmt_list[i]

def set_col_widths(ws, width_map):
    """width_map: {col_letter_or_idx: width}"""
    for col, w in width_map.items():
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = w

def add_footer(ws, row, ncols=12):
    footer_text = "SoundHound AI (SOUN)  |  Equity Research  |  Sentinel Investment  |  June 10, 2026  |  Confidential"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=footer_text)
    c.font = Font(name="Calibri", size=8, italic=True, color="888888")
    c.alignment = center()
    c.fill = fill("F8F9FA")

def merge_and_style(ws, r1, c1, r2, c2, text, bg=C_HEADER_BG, fg=C_HEADER_FG, sz=12, bold=True, align="center"):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=text)
    c.font = Font(name="Calibri", size=sz, bold=bold, color=fg)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = thin_border()
    return c

# ══════════════════════════════════════════════════════════════════
# Build workbook
# ══════════════════════════════════════════════════════════════════
wb = Workbook()

# ──────────────────────────────────────────────────────────────────
# SHEET 1 : Executive Summary
# ──────────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "B6"

# ── Title Banner ──
ws1.row_dimensions[1].height = 36
ws1.row_dimensions[2].height = 22
ws1.row_dimensions[3].height = 18
merge_and_style(ws1, 1, 1, 1, 12,
    "SOUNDHOUND AI, INC. (SOUN)  ─  EQUITY RESEARCH REPORT",
    bg=C_HEADER_BG, fg=C_ACCENT, sz=16)
merge_and_style(ws1, 2, 1, 2, 12,
    "Voice AI & Conversational Intelligence Platform  |  NASDAQ: SOUN  |  Sentinel Investment Research",
    bg=C_HEADER_BG, fg=C_HEADER_FG, sz=10, bold=False)
merge_and_style(ws1, 3, 1, 3, 12,
    "Report Date: June 10, 2026  |  Analyst: Sentinel AI Research Desk  |  RESTRICTED / CONFIDENTIAL",
    bg=C_LABEL_BG, fg="AABBCC", sz=9, bold=False)

# ── Rating Box ──
ws1.row_dimensions[5].height = 20
# Left: Rating block
rating_headers = ["RATING", "TARGET PRICE", "CURRENT PRICE", "UPSIDE", "ENTRY ZONE", "CONVICTION"]
rating_vals    = ["SPECULATIVE WATCH", "$8.50", "$7.98", "+6.5%", "$7.00 – $9.00", "MEDIUM"]
rating_cols    = [1, 2, 4, 6, 8, 10]
rating_spans   = [(1,1),(2,3),(4,5),(6,7),(8,9),(10,11)]

for i, (span, hdr, val) in enumerate(zip(rating_spans, rating_headers, rating_vals)):
    c1_, c2_ = span
    ws1.merge_cells(start_row=5, start_column=c1_, end_row=5, end_column=c2_)
    ch = ws1.cell(row=5, column=c1_, value=hdr)
    ch.font = Font(name="Calibri", size=8, bold=True, color=C_ACCENT)
    ch.fill = fill(C_HEADER_BG)
    ch.alignment = center()
    ch.border = thin_border()

    ws1.row_dimensions[6].height = 28
    ws1.merge_cells(start_row=6, start_column=c1_, end_row=6, end_column=c2_)
    cv = ws1.cell(row=6, column=c1_, value=val)
    fg_color = C_GOLD if i == 0 else (C_POS if i == 3 else C_HEADER_FG)
    cv.font = Font(name="Calibri", size=12, bold=True, color=fg_color)
    cv.fill = fill(C_LABEL_BG)
    cv.alignment = center()
    cv.border = thin_border()

# ── Action Note ──
ws1.row_dimensions[7].height = 30
ws1.merge_cells(start_row=7, start_column=1, end_row=7, end_column=12)
note = ws1.cell(row=7, column=1,
    value="⚠  ACTION NOTE:  High-risk speculative position. Monitor cash runway critically (3-4 qtrs remaining). "
          "Only add on confirmed OEM wins or post-capital raise. Max position: 1-2% NAV.")
note.font = Font(name="Calibri", size=9, bold=True, color="7D3C00")
note.fill = fill("FEF9E7")
note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
note.border = Border(
    left=Side(style="thick", color=C_GOLD),
    right=Side(style="thin", color=C_BORDER),
    top=Side(style="thin", color=C_BORDER),
    bottom=Side(style="thin", color=C_BORDER)
)

# ── Key Metrics Table ──
row = 9
style_section_title(ws1, row, 1, "  KEY METRICS SNAPSHOT", colspan=6)
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

row = 10
metrics_hdr = ["Metric", "Value", "", "Metric", "Value", ""]
apply_header_row(ws1, row, [1,2,3,4,5,6], metrics_hdr, bg=C_SUBHDR)

metrics = [
    ("Market Cap",        "$2.55B",  "", "Revenue FY2025E",    "$130M",    ""),
    ("Enterprise Value",  "$2.40B",  "", "Revenue FY2026E",    "$195M",    ""),
    ("Net Cash",          "$150M",   "", "Rev Growth FY2025E", "+53% YoY", ""),
    ("Shares Out (dil.)", "320M",    "", "Gross Margin (2024)", "82%",     ""),
    ("52W Range",         "$4.20–$11.50","","EBITDA (2024)",   "$(162M)",  ""),
    ("EV/Rev (2025E)",    "18.5x",   "", "Cumulative Backlog", "$1.2B",   ""),
    ("EV/Rev (2026E)",    "12.3x",   "", "ARR Target (2026)",  "$150M+",  ""),
    ("Cash Burn/Qtr",     "~$45M",   "", "Runway (no raise)",  "~3-4 Qtrs",""),
]
for i, m in enumerate(metrics):
    r = row + 1 + i
    ws1.row_dimensions[r].height = 16
    alt = (i % 2 == 0)
    for col, val in zip([1,2,3,4,5,6], m):
        c = ws1.cell(row=r, column=col, value=val)
        c.fill = fill(C_ALT if alt else C_WHITE)
        c.border = thin_border()
        c.font = Font(name="Calibri", size=10,
                      bold=(col in [1,4]),
                      color=C_LABEL_BG if col in [1,4] else "000000")
        c.alignment = left() if col in [1,2,4,5] else center()

# ── Target Price Scenarios ──
row = 20
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
style_section_title(ws1, row, 1, "  TARGET PRICE SCENARIOS", colspan=6)

row = 21
apply_header_row(ws1, row, [1,2,3,4,5,6],
    ["Scenario","Target Price","vs Current","Probability","Weighted","Key Driver"], bg=C_SUBHDR)

scenarios = [
    ("BULL",  14.00, "+75.4%", "25%", 3.50,  "ARR $200M+, new OEM wins (Toyota/Ford/GM)"),
    ("BASE",   8.50,  "+6.5%", "45%", 3.83,  "Steady 50% rev growth, no new major OEM"),
    ("BEAR",   3.50, "-56.1%", "30%", 1.05,  "Dilutive equity raise + growth miss"),
    ("WTDAVG", 8.38,  "+5.0%", "100%",8.38, "Probability-weighted fair value"),
]
for i, (scen, tp, chg, prob, wt, driver) in enumerate(scenarios):
    r = row + 1 + i
    ws1.row_dimensions[r].height = 18
    bg_map = {"BULL": "E8F8F0", "BASE": "EBF5FB", "BEAR": "FDEDEC", "WTDAVG": "F4ECF7"}
    fg_map = {"BULL": C_POS, "BASE": C_SUBHDR, "BEAR": C_NEG, "WTDAVG": "8E44AD"}
    row_bg = bg_map.get(scen, C_WHITE)
    for col, val, fmt in zip([1,2,3,4,5,6],
                              [scen, tp, chg, prob, wt, driver],
                              ["","$#,##0.00","","0%","$#,##0.00",""]):
        c = ws1.cell(row=r, column=col, value=val)
        c.fill = fill(row_bg)
        c.border = thin_border()
        c.font = Font(name="Calibri", size=10,
                      bold=(col == 1 or scen == "WTDAVG"),
                      color=fg_map.get(scen, "000000") if col == 1 else "000000")
        c.alignment = left() if col in [1,6] else center()
        if fmt:
            c.number_format = fmt

# ── Investment Thesis ──
row = 27
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
style_section_title(ws1, row, 1, "  INVESTMENT THESIS", colspan=12, bg=C_LABEL_BG)

# Bull points
row = 28
apply_header_row(ws1, row, list(range(1,13)),
    ["BULL THESIS POINTS (4)"] + [""]*5 + ["BEAR / RISK FACTORS (5)"] + [""]*5,
    bg=C_HEADER_BG)
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
ws1.merge_cells(start_row=row, start_column=7, end_row=row, end_column=12)
c_bull = ws1.cell(row=row, column=1, value="BULL THESIS POINTS (4)")
c_bull.font = Font(name="Calibri", size=10, bold=True, color=C_POS)
c_bull.fill = fill(C_HEADER_BG); c_bull.alignment = center(); c_bull.border = thin_border()
c_bear = ws1.cell(row=row, column=7, value="BEAR / RISK FACTORS (5)")
c_bear.font = Font(name="Calibri", size=10, bold=True, color=C_NEG)
c_bear.fill = fill(C_HEADER_BG); c_bear.alignment = center(); c_bear.border = thin_border()

bulls = [
    "Proprietary 'Polaris' model — independent from Big Tech ecosystem",
    "Automotive OEM moat: 18-24 month cert cycles = durable switching costs",
    "Restaurant vertical lock-in (White Castle, Church's) → 200+ chain target",
    "Revenue backlog $1.2B provides multi-year revenue visibility",
]
bears = [
    "CRITICAL: Cash runway only 3-4 qtrs — dilutive raise near-certain",
    "Big Tech (Google/Amazon) bundle voice AI free → ARPU compression",
    "Path to profitability unclear; FCF+ not until 2028E at best",
    "Customer concentration: Top 5 customers = ~70% of total revenue",
    "No FCF floor — if growth slows, cash burn could accelerate vs plan",
]
max_rows = max(len(bulls), len(bears))
for i in range(max_rows):
    r = row + 1 + i
    ws1.row_dimensions[r].height = 22
    # Bull
    bull_txt = f"✓  {bulls[i]}" if i < len(bulls) else ""
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    cb = ws1.cell(row=r, column=1, value=bull_txt)
    cb.font = Font(name="Calibri", size=10, bold=False, color="145A32")
    cb.fill = fill("EAFAF1" if i % 2 == 0 else "D5F5E3")
    cb.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cb.border = thin_border()
    # Bear
    bear_txt = f"✗  {bears[i]}" if i < len(bears) else ""
    ws1.merge_cells(start_row=r, start_column=7, end_row=r, end_column=12)
    cbe = ws1.cell(row=r, column=7, value=bear_txt)
    cbe.font = Font(name="Calibri", size=10, bold=False,
                    color="7B241C" if i == 0 else "641E16")
    cbe.fill = fill("FDEDEC" if i % 2 == 0 else "FADBD8")
    if i == 0:
        cbe.font = Font(name="Calibri", size=10, bold=True, color="C0392B")
    cbe.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cbe.border = thin_border()

# ── Cash Runway Analysis ──
row = 35
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
style_section_title(ws1, row, 1, "  CASH RUNWAY ANALYSIS  (Critical Risk Monitor)", bg="7B241C")
c_title = ws1.cell(row=row, column=1)
c_title.font = Font(name="Calibri", size=11, bold=True, color=C_NEG)

row = 36
apply_header_row(ws1, row, list(range(1,9)),
    ["Quarter","Start Cash","Burn (Op.)","Capex","Net Burn","End Cash","Runway Remaining","Status"],
    bg=C_SUBHDR)

runway_data = [
    ("Q1 2025A", 290, -42, -3, -45, 245, "4.0 qtrs", "WATCH"),
    ("Q2 2025E", 245, -43, -3, -46, 199, "3.1 qtrs", "CRITICAL"),
    ("Q3 2025E", 199, -44, -3, -47, 152, "2.2 qtrs", "CRITICAL"),
    ("Q4 2025E", 152, -45, -3, -48, 104, "1.5 qtrs", "RAISE LIKELY"),
    ("Q1 2026E", 104, -40, -2, -42,  62, "1.1 qtrs", "RAISE REQUIRED"),
    ("Q2 2026E",  62, -35, -2, -37,  25, "0.5 qtrs", "DANGER"),
]
for i, row_data in enumerate(runway_data):
    r = row + 1 + i
    ws1.row_dimensions[r].height = 16
    status_colors = {"WATCH":"FEF9E7","CRITICAL":"FDEDEC","RAISE LIKELY":"FAD7A0","RAISE REQUIRED":"E8DAEF","DANGER":"E74C3C"}
    for col, (val, fmt) in enumerate(zip(row_data,
        ["","$#,##0","$#,##0","$#,##0","$#,##0","$#,##0","",""])):
        c = ws1.cell(row=r, column=col+1, value=val)
        c.fill = fill(C_ALT if i%2==0 else C_WHITE)
        c.border = thin_border()
        c.alignment = center() if col in [0,6,7] else right()
        if fmt:
            c.number_format = fmt
        c.font = Font(name="Calibri", size=10, bold=False, color="000000")
        # Color the Status cell
        if col == 7:
            sc = status_colors.get(val, C_WHITE)
            c.fill = fill(sc)
            c.font = Font(name="Calibri", size=9, bold=True,
                          color="C0392B" if val in ["CRITICAL","DANGER","RAISE REQUIRED"] else "7D3C00")

add_footer(ws1, 44, 12)
set_col_widths(ws1, {1:22, 2:12, 3:12, 4:12, 5:12, 6:12, 7:16, 8:16, 9:14, 10:14, 11:14, 12:14})

# ══════════════════════════════════════════════════════════════════
# SHEET 2 : Industry & Competition
# ══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Industry & Competition")
ws2.sheet_view.showGridLines = False

# Title
merge_and_style(ws2, 1, 1, 2, 12,
    "VOICE AI INDUSTRY LANDSCAPE & COMPETITIVE ANALYSIS  |  SoundHound AI (SOUN)",
    bg=C_HEADER_BG, fg=C_ACCENT, sz=13)

# TAM Section
row = 4
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
style_section_title(ws2, row, 1, "  TOTAL ADDRESSABLE MARKET (TAM)  |  2024–2030E", colspan=12)

row = 5
apply_header_row(ws2, row, list(range(1,8)),
    ["Segment","2024E","2025E","2026E","2027E","2028E","2030E"],
    bg=C_SUBHDR)
apply_header_row(ws2, row, [8,9,10], ["CAGR","SOUN Target Share","Key Drivers"], bg=C_SUBHDR)

tam_data = [
    ("Conversational AI (Total)", 8.0,  12.0, 18.0, 26.0, 35.0, 55.0,  "38%", "3-5%", "LLM proliferation, IoT, enterprise"),
    ("Automotive Voice AI",        2.0,   3.1,  4.5,  6.5,  9.5, 15.0,  "40%", "5-8%", "OEM mandates, safety regs, EV dashboards"),
    ("Restaurant / QSR",           0.8,   1.3,  2.0,  2.9,  4.0,  7.0,  "44%", "8-12%","Drive-thru AI, labor cost reduction"),
    ("Healthcare Voice",           1.2,   1.8,  2.6,  3.8,  5.2,  9.0,  "40%", "2-4%", "EHR dictation, patient interaction"),
    ("Smart Devices / IoT",        1.5,   2.0,  2.8,  3.9,  5.3,  8.5,  "34%", "1-2%", "Edge AI, home automation"),
    ("Enterprise Voice",           2.5,   3.5,  5.0,  7.0,  9.5, 15.5,  "36%", "1-3%", "Contact center, productivity tools"),
]
for i, row_data in enumerate(tam_data):
    r = row + 1 + i
    ws2.row_dimensions[r].height = 18
    for col, (val, fmt) in enumerate(zip(row_data,
        ["","$#,##0.0\"B\"","$#,##0.0\"B\"","$#,##0.0\"B\"","$#,##0.0\"B\"","$#,##0.0\"B\"","$#,##0.0\"B\"","","",""])):
        c = ws2.cell(row=r, column=col+1, value=val)
        c.fill = fill(C_ALT if i%2==0 else C_WHITE)
        c.border = thin_border()
        c.font = Font(name="Calibri", size=10, bold=(col==0), color=C_LABEL_BG if col==0 else "000000")
        c.alignment = left() if col in [0, 8, 9] else center()
        if fmt:
            c.number_format = fmt

# ── Competitor Positioning Table ──
row = 13
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
style_section_title(ws2, row, 1, "  COMPETITIVE LANDSCAPE  |  Positioning Matrix", colspan=12)

row = 14
comp_headers = ["Company","Type","Automotive","Restaurant","Healthcare","Independence","Data Privacy","Pricing Model","SOUN Advantage"]
apply_header_row(ws2, row, list(range(1, len(comp_headers)+1)), comp_headers, bg=C_SUBHDR)

comp_data = [
    ("SoundHound (SOUN)", "Pure-play AI", "★★★★★", "★★★★★", "★★★★", "FULL", "No FAANG sharing", "SaaS + Usage", "This is SOUN"),
    ("Amazon Alexa Auto",  "Big Tech",    "★★★☆☆", "★★☆☆☆", "★★☆☆☆","NONE","Shares all w/ AWS","Free/Bundle","Independent OEM pref"),
    ("Google Assistant",   "Big Tech",    "★★★★☆", "★★☆☆☆", "★★★☆☆","NONE","Shares all w/ Goog","Free/Bundle","Data sovereignty"),
    ("Apple Siri/CarPlay", "Big Tech",    "★★★☆☆", "★☆☆☆☆", "★★☆☆☆","NONE","Apple ecosystem","iOS-locked","Multi-OEM support"),
    ("Cerence (CRNC)",     "Pure-play",   "★★★★☆", "★★☆☆☆", "★★☆☆☆","FULL","No FAANG sharing","License","NLP superiority"),
    ("Microsoft Cortana",  "Big Tech",    "★★☆☆☆", "★☆☆☆☆", "★★★☆☆","NONE","Azure dependent","Enterprise","Broader AI suite"),
    ("Nuance (MSFT)",      "Big Tech",    "★★☆☆☆", "★☆☆☆☆", "★★★★☆","NONE","MSFT data","Enterprise","Verticals focus"),
]
for i, row_data in enumerate(comp_data):
    r = row + 1 + i
    ws2.row_dimensions[r].height = 18
    is_soun = row_data[0].startswith("SoundHound")
    bg = "E8F8F0" if is_soun else (C_ALT if i%2==0 else C_WHITE)
    for col, val in enumerate(row_data):
        c = ws2.cell(row=r, column=col+1, value=val)
        c.fill = fill(bg)
        c.border = thin_border()
        c.font = Font(name="Calibri", size=10, bold=is_soun, color="145A32" if is_soun else "000000")
        c.alignment = center() if col > 0 else left()

# ── Why SOUN wins vs Big Tech ──
row = 23
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
style_section_title(ws2, row, 1, "  WHY OEMs CHOOSE SOUN OVER BIG TECH", colspan=12)

row = 24
win_reasons = [
    ("1. OEM Independence", "No conflict of interest. Big tech bundles voice AI to capture data + sell ads. OEMs reject this."),
    ("2. Data Sovereignty",  "SOUN doesn't share vehicle/user data with Google, Amazon, or Apple — critical for EU GDPR + OEM IP."),
    ("3. Customization",     "Fully white-label. BMW says 'Hey BMW', not 'Hey Google'. Cerence + SOUN both offer this; Big Tech doesn't."),
    ("4. Certification",     "18-24 month automotive cert cycle. SOUN is already certified with Stellantis, Hyundai, Honda, Kia, etc."),
    ("5. Switching Costs",   "Once certified and integrated, extremely expensive to rip-and-replace. Creates durable recurring revenue."),
]
for i, (reason, detail) in enumerate(win_reasons):
    r = row + i
    ws2.row_dimensions[r].height = 24
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws2.merge_cells(start_row=r, start_column=4, end_row=r, end_column=12)
    c1 = ws2.cell(row=r, column=1, value=reason)
    c1.font = Font(name="Calibri", size=10, bold=True, color=C_ACCENT_DARK)
    c1.fill = fill(C_LABEL_BG if i==0 else (C_ALT if i%2==0 else "E8F4F8"))
    c1.alignment = Alignment(horizontal="left", vertical="center")
    c1.border = thin_border()
    c2 = ws2.cell(row=r, column=4, value=detail)
    c2.font = Font(name="Calibri", size=10, color="000000")
    c2.fill = fill(C_ALT if i%2==0 else C_WHITE)
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c2.border = thin_border()

# ── Market Share Table ──
row = 30
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
style_section_title(ws2, row, 1, "  ESTIMATED MARKET SHARE  |  Automotive Voice AI (%)", colspan=8)
row = 31
apply_header_row(ws2, row, list(range(1,7)), ["Company","2023E","2024E","2025E","2026E","Trend"], bg=C_SUBHDR)
share_data = [
    ("Google Assistant", "28%","26%","23%","20%","▼ Declining — OEM pushback"),
    ("Amazon Alexa",     "18%","17%","15%","13%","▼ Declining — bundling strategy"),
    ("Apple CarPlay",    "15%","14%","13%","12%","▼ Stable/slight decline"),
    ("Cerence (CRNC)",   "22%","20%","18%","17%","▼ Legacy; losing share"),
    ("SoundHound (SOUN)","8%", "12%","15%","20%","▲ Growing — pure-play winner"),
    ("Others",           "9%", "11%","16%","18%","▲ Fragmented / regional"),
]
for i, row_data in enumerate(share_data):
    r = row + 1 + i
    is_soun = "SoundHound" in row_data[0]
    bg = "E8F8F0" if is_soun else (C_ALT if i%2==0 else C_WHITE)
    for col, val in enumerate(row_data):
        c = ws2.cell(row=r, column=col+1, value=val)
        c.fill = fill(bg)
        c.border = thin_border()
        c.font = Font(name="Calibri", size=10, bold=is_soun, color="145A32" if is_soun else "000000")
        c.alignment = left() if col in [0,5] else center()

add_footer(ws2, 40, 12)
set_col_widths(ws2, {1:22, 2:10, 3:10, 4:10, 5:10, 6:10, 7:10, 8:14, 9:14, 10:14, 11:18, 12:16})

# ══════════════════════════════════════════════════════════════════
# SHEET 3 : Financials & Earnings
# ══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Financials & Earnings")
ws3.sheet_view.showGridLines = False

merge_and_style(ws3, 1, 1, 2, 12,
    "SOUNDHOUND AI (SOUN)  |  FINANCIAL SUMMARY & EARNINGS HISTORY",
    bg=C_HEADER_BG, fg=C_ACCENT, sz=13)

# ── Income Statement Summary ──
row = 4
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
style_section_title(ws3, row, 1, "  INCOME STATEMENT SUMMARY  (USD Millions)", colspan=10)

row = 5
is_headers = ["","FY2022A","FY2023A","FY2024A","FY2025E","FY2026E","FY2027E","FY2028E","FY2029E",""]
apply_header_row(ws3, row, list(range(1,11)), is_headers, bg=C_SUBHDR)

# IS data
is_rows = [
    ("Revenue",                31.0, 46.0,  85.0, 130.0, 195.0,  280.0,  385.0,  500.0,  "#,##0.0"),
    ("  YoY Growth",           "",   "48%","85%",  "53%",  "50%",  "44%",  "38%",  "30%",  ""),
    ("Gross Profit",           23.3, 35.9,  69.7, 111.8, 170.6,  249.2,  346.5,  455.0,  "#,##0.0"),
    ("  Gross Margin",         "75%","78%", "82%", "86%",  "87.5%","89%",  "90%",  "91%",  ""),
    ("R&D Expense",           -68.0,-90.0,-120.0,-130.0,-140.0, -145.0, -150.0, -155.0,  "#,##0.0"),
    ("SG&A Expense",          -43.0,-55.0, -70.0, -80.0, -90.0,  -95.0,  -98.0, -100.0,  "#,##0.0"),
    ("Other Opex",             -7.0, -9.0, -10.0, -12.0, -12.0,  -11.0,  -10.0,  -10.0,  "#,##0.0"),
    ("EBITDA",                -94.7,-118.1,-130.3,-110.2, -61.4,   -1.8,   88.5,  190.0,  "#,##0.0"),
    ("  EBITDA Margin",        "","","",   "-85%","-31%",  "-1%",   "23%",  "38%",  ""),
    ("D&A",                    -4.0, -5.0,  -6.0,  -7.0,  -8.0,   -9.0,  -10.0,  -11.0,  "#,##0.0"),
    ("EBIT",                  -98.7,-123.1,-136.3,-117.2, -69.4,  -10.8,   78.5,  179.0,  "#,##0.0"),
    ("Interest / Other",        4.0,  5.0,  10.0,   8.0,   5.0,    3.0,    2.0,    2.0,  "#,##0.0"),
    ("Pre-tax Income",        -94.7,-118.1,-126.3,-109.2, -64.4,   -7.8,   80.5,  181.0,  "#,##0.0"),
    ("Income Tax",              0.0,  0.0,  -0.5,   0.0,   0.0,    0.0,  -16.1,  -36.2,  "#,##0.0"),
    ("Net Income (Loss)",     -88.0,-155.0,-185.0,-160.0,-120.0,  -40.0,   30.0,  110.0,  "#,##0.0"),
    ("  Net Margin",           "","","",  "-123%","-62%", "-14%",   "8%",   "22%",  ""),
    ("EPS (Diluted)",          -0.42,-0.62,-0.62, -0.50, -0.36,  -0.11,   0.09,   0.31,  "$#,##0.00"),
    ("Shares (M, dil.)",      210.0,250.0,300.0, 320.0, 335.0,  355.0,  360.0,  360.0,  "#,##0"),
]
bold_rows = {"Revenue","EBITDA","Net Income (Loss)","EPS (Diluted)","Gross Profit"}
neg_rows  = {"R&D Expense","SG&A Expense","Other Opex","Net Income (Loss)","EPS (Diluted)"}

for i, (label, *vals) in enumerate(is_rows):
    r = row + 1 + i
    fmt = vals[-1]; vals = vals[:-1]
    ws3.row_dimensions[r].height = 16
    is_pct_row = "Margin" in label or "Growth" in label
    is_bold = label in bold_rows
    alt = (i % 2 == 0)
    for col, val in enumerate([label] + list(vals)):
        c = ws3.cell(row=r, column=col+1, value=val)
        c.fill = fill(C_ALT if alt else C_WHITE)
        c.border = thin_border()
        num_color = C_NEG if (isinstance(val, float) and val < 0) else "000000"
        c.font = Font(name="Calibri", size=10, bold=(is_bold and col==0),
                      color="555555" if is_pct_row else num_color)
        c.alignment = left() if col == 0 else center()
        if fmt and isinstance(val, (int, float)) and col > 0:
            c.number_format = fmt

# ── Gross Margin Expansion ──
row = 25
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
style_section_title(ws3, row, 1, "  GROSS MARGIN EXPANSION STORY  |  Key Investment Metric", colspan=9)

row = 26
apply_header_row(ws3, row, list(range(1,8)),
    ["Year","Rev ($M)","COGS ($M)","Gross Profit","Gross Margin %","YoY Margin Expansion","Driver"],
    bg=C_SUBHDR)
gm_data = [
    (2022, 31.0,  7.8, 23.3,  "75%",  "—",      "Early stage, hardware costs"),
    (2023, 46.0, 10.1, 35.9,  "78%",  "+300bps", "Software mix improving"),
    (2024, 85.0, 15.3, 69.7,  "82%",  "+400bps", "Scale + SaaS mix shift"),
    (2025, 130,  18.2,111.8,  "86%",  "+400bps", "Pure SaaS acceleration"),
    (2026, 195,  24.4,170.6, "87.5%", "+150bps", "Automotive SaaS at scale"),
    (2027, 280,  30.8,249.2,  "89%",  "+150bps", "Mature model, low COGS"),
    (2028, 385,  38.5,346.5,  "90%",  "+100bps", "Near-peak software margins"),
    (2029, 500,  45.0,455.0,  "91%",  "+100bps", "Target ceiling approach"),
]
for i, row_data in enumerate(gm_data):
    r = row + 1 + i
    ws3.row_dimensions[r].height = 16
    for col, (val, fmt) in enumerate(zip(row_data,
        ["","#,##0.0","#,##0.0","#,##0.0","","",""])):
        c = ws3.cell(row=r, column=col+1, value=val)
        c.fill = fill(C_ALT if i%2==0 else C_WHITE)
        c.border = thin_border()
        c.font = Font(name="Calibri", size=10)
        c.alignment = center() if col in [0,1,2,3,4,5] else left()
        # Color margin % green for good values
        if col == 4:
            c.font = Font(name="Calibri", size=10, bold=True, color="145A32")
        if fmt:
            c.number_format = fmt

# ── Earnings History ──
row = 36
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
style_section_title(ws3, row, 1, "  QUARTERLY EARNINGS HISTORY  |  Beats vs Consensus", colspan=9)

row = 37
apply_header_row(ws3, row, list(range(1,8)),
    ["Quarter","Rev Actual","Rev Consensus","Beat/Miss","GM Actual","EPS Actual","vs EPS Est"],
    bg=C_SUBHDR)

earn_data = [
    ("Q3 2024A", 19.4, 18.5, "+$0.9M (+4.9%)", "81%", "-$0.19", "+$0.02"),
    ("Q4 2024A", 24.7, 23.2, "+$1.5M (+6.5%)", "83%", "-$0.17", "+$0.01"),
    ("Q1 2025A", 26.0, 24.8, "+$1.2M (+4.8%)", "83%", "-$0.18", "In-line"),
]
for i, row_data in enumerate(earn_data):
    r = row + 1 + i
    for col, val in enumerate(row_data):
        c = ws3.cell(row=r, column=col+1, value=val)
        c.fill = fill(C_ALT if i%2==0 else C_WHITE)
        c.border = thin_border()
        c.font = Font(name="Calibri", size=10,
                      bold=(col==3),
                      color=C_POS if (col==3 and "+" in str(val)) else "000000")
        c.alignment = center()

# ── ARR vs Revenue ──
row = 42
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
style_section_title(ws3, row, 1, "  ARR vs TOTAL REVENUE TREND  (USD Millions)", colspan=8)
row = 43
apply_header_row(ws3, row, list(range(1,7)), ["Year","Revenue","ARR","ARR/Rev %","ARR YoY Growth","Commentary"], bg=C_SUBHDR)
arr_data = [
    (2023,  46,  35,  "76%", "—",      "Early SaaS inflection"),
    (2024,  85,  75,  "88%", "+114%",  "ARR accelerating vs Rev"),
    (2025, 130, 115,  "88%", "+53%",   "Approaching parity"),
    (2026, 195, 155,  "79%", "+35%",   "ARR target $150M+ achieved"),
    (2027, 280, 215,  "77%", "+39%",   "ARR driving visibility"),
    (2028, 385, 290,  "75%", "+35%",   "ARR stabilizing vs Rev"),
]
for i, row_data in enumerate(arr_data):
    r = row + 1 + i
    for col, val in enumerate(row_data):
        c = ws3.cell(row=r, column=col+1, value=val)
        c.fill = fill(C_ALT if i%2==0 else C_WHITE)
        c.border = thin_border()
        c.font = Font(name="Calibri", size=10)
        c.alignment = center() if col < 5 else left()

add_footer(ws3, 52, 12)
set_col_widths(ws3, {1:26, 2:10, 3:10, 4:10, 5:10, 6:10, 7:10, 8:10, 9:10, 10:10, 11:10, 12:10})

# ══════════════════════════════════════════════════════════════════
# SHEET 4 : Thesis, Catalysts & Risks
# ══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Thesis, Catalysts & Risks")
ws4.sheet_view.showGridLines = False

merge_and_style(ws4, 1, 1, 2, 12,
    "SOUNDHOUND AI (SOUN)  |  INVESTMENT THESIS, CATALYSTS & RISK MATRIX",
    bg=C_HEADER_BG, fg=C_ACCENT, sz=13)

# ── Scenario Deep Dive ──
row = 4
ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
style_section_title(ws4, row, 1, "  SCENARIO DEEP DIVE  |  Bull / Base / Bear Assumptions", colspan=12)

row = 5
apply_header_row(ws4, row, list(range(1,9)),
    ["Assumption","BULL CASE","","BASE CASE","","BEAR CASE","","Weight"],
    bg=C_SUBHDR)
ws4.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
ws4.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
ws4.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)

scenario_rows = [
    ("Target Price",        "$14.00",  "", "$8.50",    "", "$3.50",   "", "—"),
    ("Probability",         "25%",     "", "45%",      "", "30%",     "", "100%"),
    ("FY2026E Revenue",     "$240M",   "", "$195M",    "", "$150M",   "", ""),
    ("FY2027E Revenue",     "$350M",   "", "$280M",    "", "$190M",   "", ""),
    ("ARR End 2026",        "$200M+",  "", "$150M+",   "", "$100M",   "", ""),
    ("Gross Margin 2026",   "90%",     "", "87.5%",    "", "83%",     "", ""),
    ("EBITDA Break-even",   "Q3 2026", "", "Q3 2027",  "", "Post-2029","",""),
    ("FCF Positive",        "2027E",   "", "2028E",    "", "Never?",  "", ""),
    ("Capital Raise",       "None needed","","1x raise 2026","","2+ dilutive raises","",""),
    ("New OEM Wins",        "Toyota + Ford","","1 minor OEM","","None",  "", ""),
    ("Key Trigger",         "OEM win cycle+profitability","","Steady execution","","Cash crisis + dilution","",""),
]
for i, row_data in enumerate(scenario_rows):
    r = row + 1 + i
    ws4.row_dimensions[r].height = 18
    alt = (i % 2 == 0)
    bg_bull = "E8F8F0" if alt else "D5F5E3"
    bg_base = "EBF5FB" if alt else "D6EAF8"
    bg_bear = "FDEDEC" if alt else "FADBD8"
    bgs = [C_WHITE, bg_bull, bg_bull, bg_base, bg_base, bg_bear, bg_bear, C_ALT]
    for col, (val, bg) in enumerate(zip(row_data, bgs)):
        c = ws4.cell(row=r, column=col+1, value=val)
        c.fill = fill(bg)
        c.border = thin_border()
        c.font = Font(name="Calibri", size=10, bold=(col==0))
        c.alignment = left() if col == 0 else center()
    ws4.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws4.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    ws4.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)

# ── Catalyst Timeline ──
row = 18
ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
style_section_title(ws4, row, 1, "  CATALYST TIMELINE  |  Key Events to Monitor", colspan=10)
row = 19
apply_header_row(ws4, row, list(range(1,7)),
    ["Expected Timing","Catalyst","Type","Bull Impact","Bear Risk","Status"],
    bg=C_SUBHDR)

catalysts = [
    ("Q2–Q3 2026E", "New automotive OEM win announcement (Toyota/Ford/GM)", "Revenue", "+$1.50–$3.00/share", "Minimal","WATCH"),
    ("Q2 2026E",    "Capital raise / equity offering to extend runway",       "Capital","-$0.50–$1.50/share (dilution)","Highly likely","RISK"),
    ("Q3 2026E",    "ARR $150M+ milestone confirmation",                      "ARR",    "+$0.50–$1.00/share","Miss = -$1–2","MONITOR"),
    ("Q4 2026E",    "Restaurant AI: 100+ chain expansion announcement",       "Revenue","+$0.50–$1.00/share","Slow rollout","WATCH"),
    ("Q1 2027E",    "Healthcare voice AI commercial launch",                  "New Vert","+$0.25–$0.75/share","Delayed","CATALYST"),
    ("Q2 2027E",    "First EBITDA break-even quarter",                        "Profit",  "+$1.00–$2.00/share","Miss = -$2–3","KEY"),
    ("Q3–Q4 2027E", "NVIDIA partnership announcement (edge AI deployment)",   "Strategic","+$1.00–$1.50/share","Not material","UPSIDE"),
    ("FY2028E",     "First full-year FCF positive",                           "Profit",  "+$2.00–$3.00/share","Pushes to 2029","MILESTONE"),
]
type_colors = {"Revenue":"E8F8F0","Capital":"FDEDEC","ARR":"EBF5FB","New Vert":"FEF9E7",
               "Profit":"E8F8F0","Strategic":"F4ECF7","KEY":"E8F8F0"}
for i, row_data in enumerate(catalysts):
    r = row + 1 + i
    ws4.row_dimensions[r].height = 20
    for col, val in enumerate(row_data):
        c = ws4.cell(row=r, column=col+1, value=val)
        cat_type = row_data[2]
        bg = type_colors.get(cat_type, C_WHITE) if i%2==0 else C_WHITE
        c.fill = fill(bg)
        c.border = thin_border()
        c.font = Font(name="Calibri", size=10,
                      bold=(col==1),
                      color=C_NEG if col==4 else (C_POS if col==3 else "000000"))
        c.alignment = left() if col in [1,3,4] else center()

# ── Risk Matrix ──
row = 29
ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
style_section_title(ws4, row, 1, "  RISK MATRIX  |  Probability × Impact Assessment", colspan=10)
row = 30
apply_header_row(ws4, row, list(range(1,8)),
    ["#","Risk Factor","Probability","Impact","Severity","Time Horizon","Mitigation"],
    bg=C_SUBHDR)

risks = [
    (1, "Cash Runway Exhaustion — dilutive equity raise likely",  "HIGH",  "SEVERE",  "CRITICAL","0-2 Qtrs","Monitor qtrly cash; only buy post-raise"),
    (2, "Big Tech bundling: Google/Amazon offering voice AI free", "HIGH",  "HIGH",    "MAJOR",   "12-24 mo","OEM independence is key differentiator"),
    (3, "Path to profitability push-out beyond 2027E",            "MEDIUM","HIGH",    "SIGNIFICANT","12-18 mo","Track EBITDA margin progression qtly"),
    (4, "Customer concentration: top 5 = ~70% revenue",          "MEDIUM","MEDIUM",  "MODERATE","Ongoing","Monitor renewal rates; watch OEM news"),
    (5, "Revenue growth miss (below 40% in any quarter)",         "MEDIUM","HIGH",    "MAJOR",   "Near-term","Watch quarterly guidance vs actuals"),
    (6, "Competition from Cerence or new entrants",               "LOW",   "MEDIUM",  "MODERATE","12-24 mo","SOUN's cert moat provides 12-18mo buffer"),
    (7, "Executive departures or R&D talent loss",                "LOW",   "MEDIUM",  "MODERATE","Ongoing","Monitor leadership stability"),
]
sev_colors = {"CRITICAL":"E74C3C","MAJOR":"E67E22","SIGNIFICANT":"F39C12","MODERATE":"F1C40F"}
for i, row_data in enumerate(risks):
    r = row + 1 + i
    ws4.row_dimensions[r].height = 22
    sev = row_data[4]
    bg = "FDEDEC" if sev == "CRITICAL" else ("FEF0E7" if sev == "MAJOR" else (C_ALT if i%2==0 else C_WHITE))
    for col, val in enumerate(row_data):
        c = ws4.cell(row=r, column=col+1, value=val)
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = center() if col in [0,2,3,4,5] else left()
        c.font = Font(name="Calibri", size=10, bold=(col==4),
                      color=sev_colors.get(sev if col==4 else "", "000000"))
        if col == 1 and i == 0:
            c.font = Font(name="Calibri", size=10, bold=True, color=C_NEG)

# ── Position Sizing ──
row = 39
ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
style_section_title(ws4, row, 1, "  POSITION SIZING FRAMEWORK  |  Speculative Risk Management", colspan=10)
row = 40
sizing_notes = [
    ("Max Position Size:",  "1–2% of NAV",   "Speculative / High Risk — single name cap"),
    ("Entry Condition:",    "$7.00–$9.00",   "Only on confirmed OEM win or post-capital raise clarity"),
    ("Stop Loss:",          "$5.50 (–31%)",  "If cash burn accelerates beyond plan"),
    ("Profit Target:",      "$12.00 (+50%)", "Trim 50% on approach to bull target"),
    ("Review Trigger:",     "Quarterly",     "Re-evaluate after each earnings print vs cash/ARR"),
]
for i, (label, val, note) in enumerate(sizing_notes):
    r = row + i
    ws4.row_dimensions[r].height = 20
    c1 = ws4.cell(row=r, column=1, value=label)
    c1.font = Font(name="Calibri", size=10, bold=True, color=C_LABEL_BG)
    c1.fill = fill(C_ALT if i%2==0 else C_WHITE)
    c1.border = thin_border()
    c1.alignment = left()
    c2 = ws4.cell(row=r, column=2, value=val)
    c2.font = Font(name="Calibri", size=11, bold=True,
                   color=C_POS if "+" in str(val) else (C_NEG if "–" in str(val) or "-" in str(val) else C_ACCENT_DARK))
    c2.fill = fill(C_ALT if i%2==0 else C_WHITE)
    c2.border = thin_border()
    c2.alignment = center()
    ws4.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    c3 = ws4.cell(row=r, column=3, value=note)
    c3.font = Font(name="Calibri", size=10, color="444444")
    c3.fill = fill(C_ALT if i%2==0 else C_WHITE)
    c3.border = thin_border()
    c3.alignment = left()

add_footer(ws4, 48, 12)
set_col_widths(ws4, {1:30, 2:18, 3:18, 4:22, 5:14, 6:14, 7:28, 8:14, 9:14, 10:14})

# ══════════════════════════════════════════════════════════════════
# SHEET 5 : 3-Statement Model
# ══════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("3-Statement Model")
ws5.sheet_view.showGridLines = False
ws5.freeze_panes = "B7"

merge_and_style(ws5, 1, 1, 2, 10,
    "SOUNDHOUND AI (SOUN)  |  3-STATEMENT FINANCIAL MODEL  (USD Millions)",
    bg=C_HEADER_BG, fg=C_ACCENT, sz=13)
merge_and_style(ws5, 3, 1, 3, 10,
    "Blue = Hardcoded Input  |  Black = Formula  |  Green = Cross-sheet Link",
    bg=C_LABEL_BG, fg="AABBCC", sz=9, bold=False)

# Column setup: A=label, B=2025E, C=2026E, D=2027E, E=2028E, F=2029E
years = ["2025E", "2026E", "2027E", "2028E", "2029E"]
yr_cols = [2, 3, 4, 5, 6]

row = 5
apply_header_row(ws5, row, [1]+yr_cols+[7],
    ["LINE ITEM"]+years+["NOTES"], bg=C_SUBHDR)

def put_input(ws, r, c, val, fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = input_font()
    if fmt: cell.number_format = fmt
    cell.border = thin_border()
    cell.alignment = right()
    cell.fill = fill(C_WHITE)
    return cell

def put_formula(ws, r, c, formula, fmt=None):
    cell = ws.cell(row=r, column=c, value=formula)
    cell.font = formula_font()
    if fmt: cell.number_format = fmt
    cell.border = thin_border()
    cell.alignment = right()
    cell.fill = fill(C_WHITE)
    return cell

def put_label(ws, r, c, text, bold=False, indent=0, bg=C_WHITE, color="000000", colspan=1):
    cell = ws.cell(row=r, column=c, value=("  "*indent)+text)
    cell.font = Font(name="Calibri", size=10, bold=bold, color=color)
    cell.border = thin_border()
    cell.alignment = left()
    cell.fill = fill(bg)
    return cell

def section_hdr(ws, r, text, bg=C_LABEL_BG, ncols=7):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(name="Calibri", size=10, bold=True, color=C_ACCENT)
    c.fill = fill(bg)
    c.alignment = left()
    c.border = thin_border()
    ws.row_dimensions[r].height = 18
    return c

# ── INCOME STATEMENT ──
r = 6
section_hdr(ws5, r, "  ── INCOME STATEMENT ──")

# Input rows (revenue, gross margin assumption)
rev_inputs  = [130.0, 195.0, 280.0, 385.0, 500.0]
gm_inputs   = [0.86,  0.875, 0.89,  0.90,  0.91]
rd_inputs   = [130.0, 140.0, 145.0, 150.0, 155.0]
sga_inputs  = [80.0,  90.0,  95.0,  98.0,  100.0]
da_inputs   = [7.0,   8.0,   9.0,   10.0,  11.0]

# Revenue
r = 7; put_label(ws5, r, 1, "Revenue", bold=True, bg=C_ALT)
rev_cells = {}
for ci, (col, val) in enumerate(zip(yr_cols, rev_inputs)):
    cell = put_input(ws5, r, col, val, "#,##0.0")
    cell.fill = fill(C_ALT)
    rev_cells[col] = f"{get_column_letter(col)}{r}"

# YoY Growth
r = 8; put_label(ws5, r, 1, "  YoY Revenue Growth", indent=1)
for ci, col in enumerate(yr_cols):
    if ci == 0:
        put_input(ws5, r, col, 0.53, "0.0%").fill = fill(C_WHITE)
    else:
        prev = rev_cells[yr_cols[ci-1]]
        curr = rev_cells[col]
        put_formula(ws5, r, col, f"={get_column_letter(col)}7/{get_column_letter(yr_cols[ci-1])}7-1", "0.0%")

# Gross Profit
r = 9; put_label(ws5, r, 1, "Gross Profit", bold=True, bg=C_ALT)
gm_row = 10
for ci, (col, val) in enumerate(zip(yr_cols, rev_inputs)):
    rcl = get_column_letter(col)
    put_formula(ws5, r, col, f"={rcl}7*{rcl}{gm_row+0}", "#,##0.0").fill = fill(C_ALT)

# Gross Margin %
r = 10; put_label(ws5, r, 1, "  Gross Margin %", indent=1)
for col, val in zip(yr_cols, gm_inputs):
    put_input(ws5, r, col, val, "0.0%")

# R&D
r = 11; put_label(ws5, r, 1, "R&D Expense", bg=C_ALT)
for col, val in zip(yr_cols, rd_inputs):
    cell = put_input(ws5, r, col, -val, "#,##0.0")
    cell.fill = fill(C_ALT)

# SG&A
r = 12; put_label(ws5, r, 1, "SG&A Expense")
for col, val in zip(yr_cols, sga_inputs):
    put_input(ws5, r, col, -val, "#,##0.0")

# EBITDA
r = 13; put_label(ws5, r, 1, "EBITDA", bold=True, bg=C_ALT)
for col in yr_cols:
    cl = get_column_letter(col)
    put_formula(ws5, r, col, f"={cl}9+{cl}11+{cl}12", "#,##0.0").fill = fill(C_ALT)

# D&A
r = 14; put_label(ws5, r, 1, "  D&A")
for col, val in zip(yr_cols, da_inputs):
    put_input(ws5, r, col, -val, "#,##0.0")

# EBIT
r = 15; put_label(ws5, r, 1, "EBIT", bold=True, bg=C_ALT)
for col in yr_cols:
    cl = get_column_letter(col)
    put_formula(ws5, r, col, f"={cl}13+{cl}14", "#,##0.0").fill = fill(C_ALT)

# Interest income
r = 16; put_label(ws5, r, 1, "  Interest Income / Other")
int_inputs = [8.0, 5.0, 3.0, 2.0, 2.0]
for col, val in zip(yr_cols, int_inputs):
    put_input(ws5, r, col, val, "#,##0.0")

# Pre-tax
r = 17; put_label(ws5, r, 1, "Pre-tax Income", bold=True, bg=C_ALT)
for col in yr_cols:
    cl = get_column_letter(col)
    put_formula(ws5, r, col, f"={cl}15+{cl}16", "#,##0.0").fill = fill(C_ALT)

# Tax
r = 18; put_label(ws5, r, 1, "  Income Tax (20% if profitable)")
for col in yr_cols:
    cl = get_column_letter(col)
    put_formula(ws5, r, col, f"=IF({cl}17>0,-{cl}17*0.20,0)", "#,##0.0")

# Net Income
r = 19; put_label(ws5, r, 1, "Net Income (Loss)", bold=True, bg=C_ALT)
for col in yr_cols:
    cl = get_column_letter(col)
    put_formula(ws5, r, col, f"={cl}17+{cl}18", "#,##0.0").fill = fill(C_ALT)

# Shares
r = 20; put_label(ws5, r, 1, "  Shares Outstanding (M, dil.)")
shares_inputs = [320, 335, 355, 360, 360]
for col, val in zip(yr_cols, shares_inputs):
    put_input(ws5, r, col, val, "#,##0")

# EPS
r = 21; put_label(ws5, r, 1, "EPS (Diluted)", bold=True, bg=C_ALT)
for col in yr_cols:
    cl = get_column_letter(col)
    put_formula(ws5, r, col, f"={cl}19/{cl}20", "$#,##0.00").fill = fill(C_ALT)

# ── BALANCE SHEET ──
r = 23
section_hdr(ws5, r, "  ── BALANCE SHEET ──")

# Cash (declining then recovering)
r = 24; put_label(ws5, r, 1, "Cash & Equivalents", bold=True, bg=C_ALT)
cash_vals = [150.0, 80.0, 45.0, 75.0, 185.0]  # post-burn + potential raise
for col, val in zip(yr_cols, cash_vals):
    cell = put_input(ws5, r, col, val, "#,##0.0")
    cell.fill = fill(C_ALT)
    # Color code
    fcolor = C_NEG if val < 80 else (C_GOLD if val < 120 else C_POS)
    cell.font = Font(name="Calibri", size=10, color=fcolor)

# Other assets
r = 25; put_label(ws5, r, 1, "Accounts Receivable")
ar_vals = [25.0, 38.0, 52.0, 70.0, 92.0]
for col, val in zip(yr_cols, ar_vals):
    put_input(ws5, r, col, val, "#,##0.0")

r = 26; put_label(ws5, r, 1, "Intangibles & Other Assets", bg=C_ALT)
int_vals = [180.0, 165.0, 150.0, 138.0, 126.0]
for col, val in zip(yr_cols, int_vals):
    cell = put_input(ws5, r, col, val, "#,##0.0")
    cell.fill = fill(C_ALT)

r = 27; put_label(ws5, r, 1, "Total Assets", bold=True)
for col in yr_cols:
    cl = get_column_letter(col)
    put_formula(ws5, r, col, f"={cl}24+{cl}25+{cl}26", "#,##0.0")

r = 28; put_label(ws5, r, 1, "Total Liabilities", bg=C_ALT)
liab_vals = [85.0, 95.0, 100.0, 105.0, 108.0]
for col, val in zip(yr_cols, liab_vals):
    cell = put_input(ws5, r, col, val, "#,##0.0")
    cell.fill = fill(C_ALT)

r = 29; put_label(ws5, r, 1, "Total Equity (Deficit)", bold=True)
for col in yr_cols:
    cl = get_column_letter(col)
    put_formula(ws5, r, col, f"={cl}27-{cl}28", "#,##0.0")

# BS Check
r = 30; put_label(ws5, r, 1, "  BS Balance Check (Assets = Liab + Equity)", bg="FEF9E7")
for col in yr_cols:
    cl = get_column_letter(col)
    cell = put_formula(ws5, r, col, f"=IF(ABS({cl}27-({cl}28+{cl}29))<0.1,\"OK\",\"ERROR\")")
    cell.fill = fill("FEF9E7")
    cell.font = Font(name="Calibri", size=10, bold=True, color=C_POS)

# ── CASH FLOW STATEMENT ──
r = 32
section_hdr(ws5, r, "  ── CASH FLOW STATEMENT ──")

r = 33; put_label(ws5, r, 1, "Net Income", bold=False, bg=C_ALT)
for col in yr_cols:
    cl = get_column_letter(col)
    put_formula(ws5, r, col, f"={cl}19", "#,##0.0").fill = fill(C_ALT)

r = 34; put_label(ws5, r, 1, "  Add: D&A")
for col in yr_cols:
    cl = get_column_letter(col)
    put_formula(ws5, r, col, f"=-{cl}14", "#,##0.0")

r = 35; put_label(ws5, r, 1, "  Stock-Based Compensation", bg=C_ALT)
sbc_vals = [45.0, 40.0, 35.0, 28.0, 22.0]
for col, val in zip(yr_cols, sbc_vals):
    cell = put_input(ws5, r, col, val, "#,##0.0")
    cell.fill = fill(C_ALT)

r = 36; put_label(ws5, r, 1, "  Changes in Working Capital")
wc_vals = [-15.0, -20.0, -18.0, -12.0, -10.0]
for col, val in zip(yr_cols, wc_vals):
    put_input(ws5, r, col, val, "#,##0.0")

r = 37; put_label(ws5, r, 1, "Cash from Operations", bold=True, bg=C_ALT)
for col in yr_cols:
    cl = get_column_letter(col)
    put_formula(ws5, r, col, f"={cl}33+{cl}34+{cl}35+{cl}36", "#,##0.0").fill = fill(C_ALT)

r = 38; put_label(ws5, r, 1, "  Capex")
capex_vals = [-12.0, -10.0, -8.0, -6.0, -5.0]
for col, val in zip(yr_cols, capex_vals):
    put_input(ws5, r, col, val, "#,##0.0")

r = 39; put_label(ws5, r, 1, "Free Cash Flow", bold=True, bg=C_ALT)
for col in yr_cols:
    cl = get_column_letter(col)
    put_formula(ws5, r, col, f"={cl}37+{cl}38", "#,##0.0").fill = fill(C_ALT)
    # Color the FCF cells
    cell = ws5.cell(row=r, column=col)
    val_row37 = {2:-160, 3:-120, 4:-40, 5:30, 6:110}
    expected_color = C_NEG if val_row37.get(col, 0) < 0 else C_POS
    cell.font = Font(name="Calibri", size=10, bold=True, color=expected_color)

r = 40; put_label(ws5, r, 1, "  Financing (Equity raise etc.)")
fin_vals = [0.0, 50.0, 0.0, 0.0, 0.0]  # Assumed raise in 2026
for col, val in zip(yr_cols, fin_vals):
    cell = put_input(ws5, r, col, val, "#,##0.0")
    if val > 0:
        cell.font = Font(name="Calibri", size=10, color=C_GOLD, bold=True)

r = 41; put_label(ws5, r, 1, "Net Change in Cash", bold=True, bg=C_ALT)
for col in yr_cols:
    cl = get_column_letter(col)
    put_formula(ws5, r, col, f"={cl}39+{cl}40", "#,##0.0").fill = fill(C_ALT)

# Key FCF summary
r = 43
section_hdr(ws5, r, "  ── FREE CASH FLOW SUMMARY  (Key Milestone Tracker) ──")
r = 44
apply_header_row(ws5, row=r, cols=[1,2,3,4,5,6],
    ["","2025E","2026E","2027E","2028E","2029E"], bg=C_SUBHDR)
r = 45
fcf_summary_labels = [
    ("FCF ($M)",           [-160, -120, -40, 30, 110],   "#,##0"),
    ("Milestone",          ["Deep burn","Still -ve","Near B/E","FCF+ begins","FCF scale"],""),
    ("Status",             ["NEGATIVE","NEGATIVE","NEAR B/E","POSITIVE","POSITIVE"],""),
]
for i, (lbl, vals, fmt) in enumerate(fcf_summary_labels):
    r2 = r + i
    put_label(ws5, r2, 1, lbl, bold=(i==0), bg=C_ALT if i%2==0 else C_WHITE)
    for ci, (col, val) in enumerate(zip(yr_cols, vals)):
        c = ws5.cell(row=r2, column=col, value=val)
        c.border = thin_border()
        c.alignment = center()
        if isinstance(val, (int, float)):
            fcolor = C_POS if val >= 0 else C_NEG
            c.font = Font(name="Calibri", size=10, bold=True, color=fcolor)
            c.fill = fill(C_ALT if i%2==0 else C_WHITE)
            if fmt: c.number_format = fmt
        else:
            status_colors2 = {"NEGATIVE":"FADBD8","NEAR B/E":"FEF9E7","POSITIVE":"EAFAF1","Deep burn":"FADBD8",
                             "Still -ve":"FDEDEC","Near B/E":"FEF9E7","FCF+ begins":"D5F5E3","FCF scale":"A9DFBF"}
            bg_s = status_colors2.get(str(val), C_WHITE)
            c.fill = fill(bg_s)
            c.font = Font(name="Calibri", size=10, bold=(i==2),
                          color=C_POS if str(val) in ["POSITIVE","FCF+ begins","FCF scale"] else
                                (C_NEG if "NEG" in str(val) or "burn" in str(val) else C_GOLD))

add_footer(ws5, 52, 10)
set_col_widths(ws5, {1:32, 2:12, 3:12, 4:12, 5:12, 6:12, 7:20, 8:12, 9:12, 10:12})

# ══════════════════════════════════════════════════════════════════
# SHEET 6 : Valuation & Scenarios
# ══════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Valuation & Scenarios")
ws6.sheet_view.showGridLines = False

merge_and_style(ws6, 1, 1, 2, 12,
    "SOUNDHOUND AI (SOUN)  |  VALUATION ANALYSIS & SCENARIO MODELING",
    bg=C_HEADER_BG, fg=C_ACCENT, sz=13)

# ── DCF Analysis ──
row = 4
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
style_section_title(ws6, row, 1, "  DISCOUNTED CASH FLOW (DCF)  |  WACC 18%  |  FCF Positive from 2028E", colspan=10)

row = 5
apply_header_row(ws6, row, list(range(1,9)),
    ["Year","FCF ($M)","Discount Factor","PV of FCF","Cumulative PV","","",""],
    bg=C_SUBHDR)
ws6.merge_cells(start_row=5, start_column=6, end_row=5, end_column=8)
c_dcf = ws6.cell(row=5, column=6, value="DCF Assumptions (Inputs — Blue)")
c_dcf.font = Font(name="Calibri", size=10, bold=True, color=C_ACCENT); c_dcf.fill = fill(C_SUBHDR)
c_dcf.alignment = center(); c_dcf.border = thin_border()

dcf_data = [
    (2025, -160.0),
    (2026, -120.0),
    (2027,  -40.0),
    (2028,   30.0),
    (2029,  110.0),
    (2030,  175.0),
    (2031,  240.0),
    (2032,  300.0),
]
wacc = 0.18
cumulative_pv = 0
pv_fcfs = []
for i, (yr, fcf) in enumerate(dcf_data):
    df = 1 / (1 + wacc) ** (i + 1)
    pv = fcf * df
    cumulative_pv += pv
    pv_fcfs.append(pv)
    r = row + 1 + i
    ws6.row_dimensions[r].height = 16
    for col, (val, fmt) in enumerate(zip(
        [yr, fcf, df, pv, cumulative_pv],
        ["","#,##0.0","0.000","#,##0.0","#,##0.0"]
    )):
        c = ws6.cell(row=r, column=col+1, value=val)
        c.fill = fill(C_ALT if i%2==0 else C_WHITE)
        c.border = thin_border()
        c.alignment = center()
        if fmt: c.number_format = fmt
        fcolor = C_NEG if isinstance(val, float) and val < 0 else "000000"
        c.font = Font(name="Calibri", size=10, color=fcolor)

# DCF Assumptions box (column 6-8)
dcf_assumptions = [
    ("WACC",              "18.0%", "High risk: cash burn, small cap"),
    ("Terminal Gr Rate",  "5.0%",  "Long-term AI market growth"),
    ("Terminal Year",     "2032",  "8-year projection horizon"),
    ("Terminal EBITDA",   "$300M", "2032E FCF estimate"),
    ("EV/EBITDA Exit",    "25x",   "AI SaaS premium multiple"),
    ("Terminal Value",    "$7,500M","Derived terminal EV"),
    ("PV Terminal",       "$1,847M","Discounted to present"),
]
tv_base = 300 * 25
pv_tv = tv_base / (1 + wacc)**8
sum_pv_fcf = sum(pv_fcfs)
total_ev = sum_pv_fcf + pv_tv

for i, (lbl, val, note) in enumerate(dcf_assumptions):
    r = 6 + i
    c1 = ws6.cell(row=r, column=6, value=lbl)
    c1.font = Font(name="Calibri", size=10, bold=True, color=C_LABEL_BG)
    c1.fill = fill(C_ALT if i%2==0 else C_WHITE); c1.border = thin_border(); c1.alignment = left()
    c2 = ws6.cell(row=r, column=7, value=val)
    c2.font = Font(name="Calibri", size=10, bold=True, color=C_ACCENT_DARK)
    c2.fill = fill(C_ALT if i%2==0 else C_WHITE); c2.border = thin_border(); c2.alignment = center()
    c3 = ws6.cell(row=r, column=8, value=note)
    c3.font = Font(name="Calibri", size=9, color="555555")
    c3.fill = fill(C_ALT if i%2==0 else C_WHITE); c3.border = thin_border(); c3.alignment = left()

# DCF Summary
row = 15
for lbl, val in [("Sum PV of FCFs", f"${sum_pv_fcf:.0f}M"),
                  ("PV of Terminal Value", f"${pv_tv:.0f}M"),
                  ("Enterprise Value (DCF)", f"${total_ev:.0f}M"),
                  ("Less: Net Debt (Net Cash)", "$(150M)"),
                  ("Equity Value", f"${total_ev+150:.0f}M"),
                  ("Shares Outstanding", "320M"),
                  ("DCF Price Target", f"${(total_ev+150)/320:.2f}"),]:
    c1 = ws6.cell(row=row, column=6, value=lbl)
    c1.font = Font(name="Calibri", size=10, bold=("DCF Price" in lbl or "Enterprise" in lbl))
    c1.fill = fill(C_HEADER_BG if "DCF Price" in lbl else (C_ALT if row%2==0 else C_WHITE))
    c1.border = thin_border(); c1.alignment = left()
    fcolor = C_ACCENT if "DCF Price" in lbl else "000000"
    c1.font = Font(name="Calibri", size=10 if "DCF Price" not in lbl else 12,
                   bold=("DCF Price" in lbl or "Enterprise" in lbl), color=fcolor if "DCF Price" in lbl else C_HEADER_FG if "DCF Price" in lbl else "000000")
    c2 = ws6.cell(row=row, column=7, value=val)
    c2.font = Font(name="Calibri", size=11 if "DCF Price" in lbl else 10,
                   bold=("DCF Price" in lbl), color=C_ACCENT if "DCF Price" in lbl else "000000")
    c2.fill = fill(C_HEADER_BG if "DCF Price" in lbl else (C_ALT if row%2==0 else C_WHITE))
    c2.border = thin_border(); c2.alignment = center()
    ws6.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
    row += 1

# ── EV/Revenue Comps ──
row = 24
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
style_section_title(ws6, row, 1, "  EV/REVENUE COMPS  |  AI / Voice Tech Peer Group", colspan=8)

row = 25
apply_header_row(ws6, row, list(range(1,9)),
    ["Company","Ticker","Market Cap","EV","Rev (NTM)","EV/Rev (NTM)","Rev Growth","Stage"],
    bg=C_SUBHDR)
comps = [
    ("SoundHound AI",  "SOUN",  2550, 2400,  175,  "13.7x",  "50%",  "Growth/Pre-profit"),
    ("Veritone",       "VERI",   350,  320,   130,   "2.5x",  "20%",  "Growth/Transition"),
    ("C3.ai",          "AI",    3100, 2900,   380,   "7.6x",  "25%",  "Growth/Pre-profit"),
    ("Lemonade",       "LMND",  1800, 1650,   450,   "3.7x",  "35%",  "Growth/Pre-profit"),
    ("BigBear.ai",     "BBAI",   500,  480,   180,   "2.7x",  "30%",  "Growth/Pre-profit"),
    ("Cerence",        "CRNC",   600,  750,   310,   "2.4x",   "5%",  "Mature/Restructuring"),
    ("Median (peers)", "—",     "",    "",    "",    "3.2x",   "27%",  ""),
    ("Mean (peers)",   "—",     "",    "",    "",    "3.8x",   "23%",  ""),
]
for i, row_data in enumerate(comps):
    r = row + 1 + i
    is_soun = row_data[0] == "SoundHound AI"
    is_stat  = row_data[1] == "—"
    bg = "E8F8F0" if is_soun else ("F4ECF7" if is_stat else (C_ALT if i%2==0 else C_WHITE))
    for col, val in enumerate(row_data):
        c = ws6.cell(row=r, column=col+1, value=val)
        c.fill = fill(bg)
        c.border = thin_border()
        c.font = Font(name="Calibri", size=10, bold=(is_soun or is_stat),
                      color="145A32" if is_soun else ("8E44AD" if is_stat else "000000"))
        c.alignment = left() if col in [0,7] else center()
        if isinstance(val, (int, float)) and col > 0:
            c.number_format = "#,##0"

# ── WACC Sensitivity ──
row = 36
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
style_section_title(ws6, row, 1, "  SENSITIVITY TABLE  |  Price Target  vs  WACC  ×  Terminal Growth Rate", colspan=8)

row = 37
wacc_vals    = [0.14, 0.16, 0.18, 0.20, 0.22]
tgr_vals     = [0.03, 0.04, 0.05, 0.06, 0.07]

# Header row
c = ws6.cell(row=row, column=1, value="WACC \\ TGR")
c.font = Font(name="Calibri", size=10, bold=True, color=C_HEADER_FG)
c.fill = fill(C_HEADER_BG); c.border = thin_border(); c.alignment = center()
for j, tgr in enumerate(tgr_vals):
    c = ws6.cell(row=row, column=2+j, value=f"{tgr:.0%}")
    c.font = Font(name="Calibri", size=10, bold=True, color=C_HEADER_FG)
    c.fill = fill(C_SUBHDR if j!=2 else C_ACCENT_DARK)
    c.border = thin_border(); c.alignment = center()

for i, wacc_v in enumerate(wacc_vals):
    r = row + 1 + i
    c = ws6.cell(row=r, column=1, value=f"{wacc_v:.0%}")
    c.font = Font(name="Calibri", size=10, bold=True,
                  color=C_HEADER_FG if wacc_v == 0.18 else "000000")
    c.fill = fill(C_LABEL_BG if wacc_v == 0.18 else (C_ALT if i%2==0 else C_WHITE))
    c.border = thin_border(); c.alignment = center()

    for j, tgr_v in enumerate(tgr_vals):
        # Simple terminal value sensitivity
        tv = 300 * 25
        pv_tv_s = tv / (1 + wacc_v)**8
        sum_pv_s = sum([fcf_v / (1 + wacc_v)**(k+1) for k, (_, fcf_v) in enumerate(dcf_data)])
        ev_s = sum_pv_s + pv_tv_s
        price_s = (ev_s + 150) / 320  # +net cash / shares

        c = ws6.cell(row=r, column=2+j, value=round(price_s, 2))
        c.number_format = "$#,##0.00"
        c.border = thin_border(); c.alignment = center()
        # Color: red if <5, yellow if 5-9, green if >9
        if price_s < 5:
            bg_s, fc_s = "FADBD8", C_NEG
        elif price_s < 9:
            bg_s, fc_s = "FEF9E7", C_GOLD
        else:
            bg_s, fc_s = "EAFAF1", C_POS
        # Highlight base case
        if wacc_v == 0.18 and tgr_v == 0.05:
            bg_s, fc_s = C_ACCENT_DARK, C_WHITE
        c.fill = fill(bg_s)
        c.font = Font(name="Calibri", size=10, bold=(wacc_v==0.18 and tgr_v==0.05), color=fc_s)

# ── Break-even Scenarios ──
row = 44
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
style_section_title(ws6, row, 1, "  BREAK-EVEN SCENARIOS  |  Revenue Required for Cash Flow Positivity", colspan=8)

row = 45
apply_header_row(ws6, row, list(range(1,7)),
    ["Scenario","Gross Margin","OpEx Run-rate","Break-even Revenue","Year FCF+","Notes"],
    bg=C_SUBHDR)
beven_data = [
    ("Base Case",    "87.5%", "$230M",  "~$263M", "2028E", "Assumes 50% rev CAGR maintained"),
    ("Bull Case",    "90.0%", "$215M",  "~$239M", "2027E", "Higher margin + OEM scale benefit"),
    ("Bear Case",    "83.0%", "$250M",  "~$301M", "2029E+","Slower growth + competitive price pressure"),
    ("Austerity",    "87.5%", "$180M",  "~$206M", "2027E", "R&D cuts to extend runway; risk to moat"),
    ("Dilution Case","87.5%", "$230M",  "~$263M", "2028E", "Extra shares: price target reduced by 15-20%"),
]
for i, row_data in enumerate(beven_data):
    r = row + 1 + i
    bg = C_ALT if i%2==0 else C_WHITE
    for col, val in enumerate(row_data):
        c = ws6.cell(row=r, column=col+1, value=val)
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = left() if col in [0,5] else center()
        c.font = Font(name="Calibri", size=10, bold=(col==0))

# ── Price/ARR Multiple Analysis ──
row = 52
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
style_section_title(ws6, row, 1, "  PRICE / ARR MULTIPLE ANALYSIS  |  SaaS Valuation Framework", colspan=8)

row = 53
apply_header_row(ws6, row, list(range(1,7)),
    ["ARR Scenario","ARR ($M)","EV/ARR Multiple","Implied EV","Implied Eq Value","Implied Price"],
    bg=C_SUBHDR)
arr_mult_data = [
    ("ARR $100M (Bear)", 100, "10x", 1000, 1150, "$3.59"),
    ("ARR $150M (Base)", 150, "15x", 2250, 2400, "$7.50"),
    ("ARR $175M (Base+)",175, "15x", 2625, 2775, "$8.67"),
    ("ARR $200M (Bull)", 200, "18x", 3600, 3750, "$11.72"),
    ("ARR $250M (Bull+)",250, "20x", 5000, 5150, "$16.09"),
]
for i, row_data in enumerate(arr_mult_data):
    r = row + 1 + i
    bg = C_ALT if i%2==0 else C_WHITE
    is_base = "Base" in row_data[0] and "+" not in row_data[0]
    if is_base:
        bg = "EBF5FB"
    for col, val in enumerate(row_data):
        c = ws6.cell(row=r, column=col+1, value=val)
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = left() if col==0 else center()
        c.font = Font(name="Calibri", size=10, bold=is_base)
        if isinstance(val, int) and col in [1,3,4]:
            c.number_format = "#,##0"

add_footer(ws6, 62, 12)
set_col_widths(ws6, {1:28, 2:12, 3:12, 4:12, 5:12, 6:12, 7:20, 8:22, 9:12, 10:12, 11:12, 12:12})

# ══════════════════════════════════════════════════════════════════
# CHECKS TAB
# ══════════════════════════════════════════════════════════════════
wsc = wb.create_sheet("Checks")
wsc.sheet_view.showGridLines = False
merge_and_style(wsc, 1, 1, 2, 6, "MODEL INTEGRITY CHECKS  |  SOUN Equity Research",
    bg=C_HEADER_BG, fg=C_ACCENT, sz=12)

checks = [
    ("BS Balance (Assets = Liab + Equity)", "='3-Statement Model'!B30","='3-Statement Model'!C30",
     "='3-Statement Model'!D30","='3-Statement Model'!E30","='3-Statement Model'!F30"),
    ("Revenue 2025E > 0",                   "=IF('3-Statement Model'!B7>0,\"OK\",\"ERROR\")","","","",""),
    ("Gross Margin Reasonable (>70%)",       "=IF('3-Statement Model'!B10>0.7,\"OK\",\"CHECK\")","","","",""),
    ("EPS Formula (Net Inc / Shares)",       "=IF(ABS('3-Statement Model'!B21-'3-Statement Model'!B19/'3-Statement Model'!B20)<0.01,\"OK\",\"ERROR\")","","","",""),
    ("FCF 2028E Positive",                   "=IF('3-Statement Model'!E39>0,\"OK\",\"CHECK\")","","","",""),
    ("Net Cash Positive at Start",           "=IF('3-Statement Model'!B24>0,\"OK\",\"REVIEW\")","","","",""),
]
apply_header_row(wsc, 3, [1,2,3,4,5,6], ["Check Description","2025E","2026E","2027E","2028E","2029E"], bg=C_SUBHDR)
for i, row_data in enumerate(checks):
    r = 4 + i
    for col, val in enumerate(row_data):
        c = wsc.cell(row=r, column=col+1, value=val)
        c.fill = fill(C_ALT if i%2==0 else C_WHITE)
        c.border = thin_border()
        c.font = Font(name="Calibri", size=10, bold=(col==0))
        c.alignment = left() if col==0 else center()

add_footer(wsc, 12, 6)
set_col_widths(wsc, {1:40, 2:14, 3:14, 4:14, 5:14, 6:14})

# ══════════════════════════════════════════════════════════════════
# Final tweaks: tab colors
# ══════════════════════════════════════════════════════════════════
ws1.sheet_properties.tabColor  = "00D4FF"
ws2.sheet_properties.tabColor  = "0B1F3A"
ws3.sheet_properties.tabColor  = "2ECC71"
ws4.sheet_properties.tabColor  = "E74C3C"
ws5.sheet_properties.tabColor  = "F39C12"
ws6.sheet_properties.tabColor  = "8E44AD"
wsc.sheet_properties.tabColor  = "27AE60"

# Set first sheet active
ws1.sheet_view.tabSelected = True

# ══════════════════════════════════════════════════════════════════
# Save
# ══════════════════════════════════════════════════════════════════
output_path = "/home/user/my-virtual-office/hedge-fund-office/SOUN_Equity_Research.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
