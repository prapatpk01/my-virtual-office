"""CRWV Equity Research Workbook - Part 1: Setup + Sheet 1 (Executive Summary) + Sheet 2 (Industry)"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.series import SeriesLabel

# ── Colour palette ──────────────────────────────────────────────────────────
NAVY      = "0D1B2A"
NAVY2     = "1E3A5F"
GOLD      = "D4AF37"
WHITE     = "FFFFFF"
GREEN     = "2ECC71"
RED       = "E74C3C"
ORANGE    = "F39C12"
ROW_ALT   = "F0F4F8"
BORDER_C  = "CCCCCC"
BLUE_IN   = "0000FF"   # hardcoded input font
LIGHT_BLUE= "DCE6F1"
LIGHT_GREEN="E8F5E9"
LIGHT_RED  ="FDECEA"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def border(style="thin", color=BORDER_C):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_cell(ws, row, col, value, fnt=None, fill_=None, brd=None, aln=None,
             num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if fnt:   c.font      = fnt
    if fill_: c.fill      = fill_
    if brd:   c.border    = brd
    if aln:   c.alignment = aln
    if num_fmt: c.number_format = num_fmt
    return c

def merge_header(ws, row, c1, c2, text, bg=NAVY, fg=GOLD, sz=12, bold=True):
    ws.merge_cells(start_row=row, start_column=c1,
                   end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=text)
    c.font      = Font(bold=bold, color=fg, size=sz, name="Calibri")
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    return c

def footer_row(ws, row, ncols=8):
    txt = "CoreWeave (CRWV) | Equity Research | Sentinel Investment | June 10, 2026 | Confidential"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=txt)
    c.font      = Font(italic=True, color="888888", size=8, name="Calibri")
    c.fill      = fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")

def thin_brd():
    return border("thin", BORDER_C)

def section_title(ws, row, c1, c2, text):
    ws.merge_cells(start_row=row, start_column=c1,
                   end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=text)
    c.font      = Font(bold=True, color=GOLD, size=11, name="Calibri")
    c.fill      = fill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")

# ════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ════════════════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)   # remove default sheet

# ── SHEET 1: Executive Summary ───────────────────────────────────────────────
ws1 = wb.create_sheet("Executive Summary")
ws1.sheet_properties.tabColor = NAVY

# Row heights
ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 20
ws1.row_dimensions[3].height = 20

# --- Section A: Company Header ---
merge_header(ws1, 1, 1, 8, "CRWV — CoreWeave, Inc.",
             bg=NAVY, fg=GOLD, sz=16, bold=True)

ws1.merge_cells("A2:H2")
c2 = ws1["A2"]
c2.value     = "Equity Research  |  AI Cloud Infrastructure  |  June 10, 2026"
c2.font      = Font(italic=True, color="AAAAAA", size=10, name="Calibri")
c2.fill      = fill(NAVY)
c2.alignment = Alignment(horizontal="center", vertical="center")

ws1.merge_cells("A3:H3")
c3 = ws1["A3"]
c3.value     = "Rating: BUY  |  Target Price: $125  |  Current: $102.37  |  Upside: +22.1%"
c3.font      = Font(bold=True, color=GREEN, size=11, name="Calibri")
c3.fill      = fill(NAVY2)
c3.alignment = Alignment(horizontal="center", vertical="center")

ws1.row_dimensions[4].height = 6  # spacer

# --- Section B: Key Metrics header ---
section_title(ws1, 5, 1, 4, "  KEY METRICS")

col_hdrs = ["Metric", "Value", "Metric", "Value"]
for ci, hdr in enumerate(col_hdrs, 1):
    c = ws1.cell(row=6, column=ci, value=hdr)
    c.font      = Font(bold=True, color=WHITE, size=10, name="Calibri")
    c.fill      = fill(NAVY2)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin_brd()

metrics = [
    ("Current Price", "$102.37",      "Market Cap",        "$42.5B"),
    ("52W High",      "$134.80",      "52W Low",           "$38.15"),
    ("EV",            "$50.0B",       "Net Debt",          "$7.5B"),
    ("2024A Revenue", "$1,915M",      "2025E Revenue",     "$4,200M"),
    ("Revenue Growth YoY", "+119% est","Gross Margin",     "~52%"),
    ("EV/Revenue (2025E)", "11.9x",   "EV/Revenue (2026E)","6.7x"),
    ("Backlog",       "$23B+",        "Customer #1",       "Microsoft 62%"),
    ("IPO Price",     "$40.00",       "IPO Date",          "Mar 28, 2025"),
]
for i, (m1, v1, m2, v2) in enumerate(metrics):
    row = 7 + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    for ci, val in enumerate([m1, v1, m2, v2], 1):
        bold = ci in (1, 3)
        c = ws1.cell(row=row, column=ci, value=val)
        c.font      = Font(bold=bold, size=10, name="Calibri")
        c.fill      = fill(bg)
        c.border    = thin_brd()
        c.alignment = Alignment(horizontal="center" if ci % 2 == 0 else "left",
                                vertical="center")

# --- Section C: Target Price Summary ---
ws1.row_dimensions[16].height = 6
section_title(ws1, 17, 1, 5, "  TARGET PRICE SUMMARY")

tp_hdrs = ["Scenario", "Price Target", "Upside/Downside", "Probability", "Key Driver"]
for ci, hdr in enumerate(tp_hdrs, 1):
    c = ws1.cell(row=18, column=ci, value=hdr)
    c.font      = Font(bold=True, color=WHITE, size=10, name="Calibri")
    c.fill      = fill(NAVY2)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin_brd()

tp_data = [
    ("Bull",     "$180",  "+75.8%", "30%", "Hyperscaler contract expansion + margin inflection", GREEN),
    ("Base",     "$125",  "+22.1%", "50%", "Steady backlog execution, FCF positive 2027",        ORANGE),
    ("Bear",     "$55",   "-46.3%", "20%", "Customer concentration materializes, capex overrun", RED),
    ("Weighted", "$127",  "+24.1%", "—",   "Probability-weighted",                               NAVY2),
]
for i, (sc, pt, ud, prob, drv, color) in enumerate(tp_data):
    row = 19 + i
    for ci, val in enumerate([sc, pt, ud, prob, drv], 1):
        c = ws1.cell(row=row, column=ci, value=val)
        c.font   = Font(bold=(ci == 1), color=WHITE if color == NAVY2 else "000000",
                        size=10, name="Calibri")
        c.fill   = fill(color) if ci == 1 else fill(ROW_ALT if i % 2 == 0 else WHITE)
        c.border = thin_brd()
        c.alignment = Alignment(horizontal="center" if ci != 5 else "left",
                                vertical="center", wrap_text=(ci == 5))

# --- Section D: Investment Thesis ---
ws1.row_dimensions[24].height = 6
section_title(ws1, 25, 1, 8, "  INVESTMENT THESIS")

bull_header = ws1.cell(row=26, column=1, value="BULL CASE — Key Positives")
bull_header.font  = Font(bold=True, color=GREEN, size=10, name="Calibri")
bull_header.fill  = fill(ROW_ALT)
bull_header.alignment = Alignment(horizontal="left", vertical="center")
ws1.merge_cells("A26:H26")

bulls = [
    "►  Pure-play GPU cloud with $23B+ contracted backlog providing revenue visibility",
    "►  NVIDIA strategic partnership ensures priority GPU allocation (H100/H200/Blackwell)",
    "►  119% YoY revenue growth trajectory in $600B AI cloud TAM by 2030",
    "►  Gross margin expansion path from 47% (2024) → 59% (2029E)",
    "►  Microsoft partnership anchors revenue + validates enterprise-grade infrastructure",
]
for i, txt in enumerate(bulls):
    row = 27 + i
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws1.cell(row=row, column=1, value=txt)
    c.font      = Font(color="1A7A3F", size=10, name="Calibri")
    c.fill      = fill("E8F8EF")
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = border("thin", "BBBBBB")

bear_header = ws1.cell(row=33, column=1, value="BEAR CASE — Key Risks")
bear_header.font  = Font(bold=True, color=RED, size=10, name="Calibri")
bear_header.fill  = fill(ROW_ALT)
bear_header.alignment = Alignment(horizontal="left", vertical="center")
ws1.merge_cells("A33:H33")

bears = [
    "◄  Customer concentration: Microsoft = 62% of revenue (single-point failure risk)",
    "◄  $9B gross debt → heavy interest burden delaying profitability",
    "◄  Hyperscalers (AWS, Azure, GCP) building own GPU infrastructure",
    "◄  AI spending slowdown could compress valuations across the sector",
]
for i, txt in enumerate(bears):
    row = 34 + i
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws1.cell(row=row, column=1, value=txt)
    c.font      = Font(color="922B21", size=10, name="Calibri")
    c.fill      = fill("FDECEA")
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = border("thin", "BBBBBB")

footer_row(ws1, 40, 8)

# freeze + col widths
ws1.freeze_panes = "A4"
ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 22
ws1.column_dimensions["D"].width = 16
ws1.column_dimensions["E"].width = 45
for col in ["F","G","H"]:
    ws1.column_dimensions[col].width = 12

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2: Industry & Competition
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Industry & Competition")
ws2.sheet_properties.tabColor = "1E3A5F"

merge_header(ws2, 1, 1, 7,
             "AI Cloud Infrastructure — Total Addressable Market & Competitive Landscape",
             bg=NAVY, fg=GOLD, sz=13)
ws2.row_dimensions[1].height = 28

# --- Section A: TAM Table ---
section_title(ws2, 2, 1, 7, "  SECTION A — TAM & Growth (All figures in $B unless noted)")

tam_hdrs = ["Year", "TAM ($B)", "CAGR YoY", "CRWV Est. Share", "CRWV Rev ($M)", "", ""]
for ci, h in enumerate(tam_hdrs[:5], 1):
    c = ws2.cell(row=3, column=ci, value=h)
    c.font      = Font(bold=True, color=WHITE, size=10, name="Calibri")
    c.fill      = fill(NAVY2)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin_brd()

tam_data = [
    ("2022",  15,  "—",    "0.1%",  16),
    ("2023",  28,  "87%",  "0.8%",  229),
    ("2024",  65,  "132%", "2.9%",  1915),
    ("2025E", 120, "85%",  "3.5%",  4200),
    ("2026E", 200, "67%",  "3.8%",  7500),
    ("2027E", 300, "50%",  "3.7%",  11000),
    ("2028E", 400, "33%",  "3.6%",  14500),
    ("2030E", 580, "20%",  "3.2%",  18500),
]
for i, row_data in enumerate(tam_data):
    row = 4 + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    for ci, val in enumerate(row_data, 1):
        c = ws2.cell(row=row, column=ci, value=val)
        c.font      = Font(size=10, name="Calibri",
                          bold=(ci==1),
                          color=NAVY if ci==1 else "000000")
        c.fill      = fill(bg)
        c.border    = thin_brd()
        c.alignment = Alignment(horizontal="center", vertical="center")

# TAM CAGR note
ws2.merge_cells("A13:E13")
note = ws2["A13"]
note.value = "Overall TAM CAGR 2024-2030: ~44%  |  Source: Internal estimates, IDC, Gartner"
note.font  = Font(italic=True, color="555555", size=9, name="Calibri")
note.fill  = fill(ROW_ALT)

# --- Bar Chart: TAM Growth ---
chart = BarChart()
chart.type = "col"
chart.title = "AI Cloud TAM ($B) — 2022 to 2030E"
chart.style = 10
chart.y_axis.title = "TAM ($B)"
chart.x_axis.title = "Year"
chart.width = 18
chart.height = 12

data_ref   = Reference(ws2, min_col=2, min_row=3, max_row=11)
cats_ref   = Reference(ws2, min_col=1, min_row=4, max_row=11)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.solidFill = "1E3A5F"
ws2.add_chart(chart, "G2")

# --- Section B: Competitor Comparison ---
section_title(ws2, 16, 1, 7, "  SECTION B — Competitor Landscape")

comp_hdrs = ["Company", "Rev 2024", "Rev Growth", "Gross Margin", "EV", "EV/Rev", "GPU Specialization"]
for ci, h in enumerate(comp_hdrs, 1):
    c = ws2.cell(row=17, column=ci, value=h)
    c.font      = Font(bold=True, color=WHITE, size=10, name="Calibri")
    c.fill      = fill(NAVY2)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin_brd()

comp_data = [
    ("CoreWeave (CRWV)", "$1.9B",  "+728%",     "47%",   "$50B",    "26x",   "Pure-play GPU ★"),
    ("AWS",              "$107B",  "+17%",      "~60%",  "—",       "—",     "General cloud"),
    ("Azure (MSFT)",     "$135B",  "+21%",      "~68%",  "—",       "—",     "General + AI"),
    ("GCP (GOOGL)",      "$43B",   "+30%",      "~55%",  "—",       "—",     "General + AI"),
    ("Lambda Labs",      "~$0.8B", "+200% est", "~40%",  "Private", "—",     "GPU Cloud"),
    ("Oracle Cloud",     "$19.8B", "+25%",      "~58%",  "—",       "—",     "Growing AI infra"),
]
for i, row_data in enumerate(comp_data):
    row = 18 + i
    bg = LIGHT_BLUE if i == 0 else (ROW_ALT if i % 2 == 0 else WHITE)
    for ci, val in enumerate(row_data, 1):
        c = ws2.cell(row=row, column=ci, value=val)
        c.font      = Font(bold=(i==0), size=10, name="Calibri",
                          color=NAVY if i==0 else "000000")
        c.fill      = fill(bg)
        c.border    = thin_brd()
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                vertical="center")

# --- Section C: Moat Analysis ---
section_title(ws2, 26, 1, 6, "  SECTION C — Competitive Moat Analysis")

moat_hdrs = ["Moat Factor", "CRWV", "AWS", "Azure", "GCP", "CRWV Score"]
for ci, h in enumerate(moat_hdrs, 1):
    c = ws2.cell(row=27, column=ci, value=h)
    c.font      = Font(bold=True, color=WHITE, size=10, name="Calibri")
    c.fill      = fill(NAVY2)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin_brd()

moat_data = [
    ("GPU First-Mover",      "★★★★★", "★★★", "★★★", "★★★",  "High",        GREEN),
    ("NVIDIA Partnership",   "★★★★★", "★★★", "★★★", "★★★★", "High",        GREEN),
    ("Contract Lock-in",     "★★★★",  "★★★", "★★★", "★★★",  "High",        GREEN),
    ("Switching Costs",      "★★★★",  "★★★★","★★★★","★★★★", "Medium-High", ORANGE),
    ("Brand/Scale",          "★★★",   "★★★★★","★★★★★","★★★★★","Low",        RED),
    ("Price Competitiveness","★★★",   "★★★★","★★★★","★★★",  "Medium",      ORANGE),
]
for i, (factor, crwv, aws, azure, gcp, score, sc) in enumerate(moat_data):
    row = 28 + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    for ci, val in enumerate([factor, crwv, aws, azure, gcp, score], 1):
        c = ws2.cell(row=row, column=ci, value=val)
        c.font   = Font(size=10, name="Calibri",
                       color=(WHITE if ci==6 else "000000"),
                       bold=(ci==1 or ci==6))
        c.fill   = fill(sc) if ci==6 else fill(bg)
        c.border = thin_brd()
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                vertical="center")

footer_row(ws2, 37, 7)

ws2.freeze_panes = "A4"
ws2.column_dimensions["A"].width = 22
for col in ["B","C","D","E","F"]:
    ws2.column_dimensions[col].width = 15
ws2.column_dimensions["G"].width = 20

# Save progress
out_path = "/home/user/my-virtual-office/hedge-fund-office/CRWV_Equity_Research.xlsx"
wb.save(out_path)
print(f"Part 1 saved: {out_path}")
