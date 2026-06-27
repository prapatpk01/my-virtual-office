"""
Detailed backtest export to XLSX with full trade-by-trade logs.

Compares the two live-relevant configs on real Jan-May 2026 data:
  A  Current FAST v2 (LIVE NOW) — composite bias>60, pull±2.5%, classic 4-score,
                                  ADX 18-50 rising, cooldown 4, TP1=0.5R(40%)→BE, TP2=3R
  L  SJ Hybrid (PROPOSED)       — same gates but 5-component SJ score
                                  (HMA20 + EMA5>SMA9 + RSI + Volume + Breakout(10), min 4/5)

Output: /home/user/my-virtual-office/SJ_Fast_Entry_Backtest.xlsx
  Sheet 1  Summary           — side-by-side comparison
  Sheet 2  Trades_A_BTC      — every BTC trade under Current FAST v2
  Sheet 3  Trades_A_XAU
  Sheet 4  Trades_L_BTC      — every BTC trade under SJ Hybrid
  Sheet 5  Trades_L_XAU
  Sheet 6  Monthly           — month-by-month PnL breakdown
"""
import numpy as np, pandas as pd, warnings, glob, os
warnings.filterwarnings("ignore")

DATA_ROOT = "/home/user/my-virtual-office/data"
OUT_PATH  = "/home/user/my-virtual-office/SJ_Fast_Entry_Backtest.xlsx"
COLS = ["open_time","open","high","low","close","volume",
        "close_time","quote_volume","count","tbv","tbqv","ignore"]

def load_tf(sym, tf):
    folder = os.path.join(DATA_ROOT, "btc_csv" if sym == "BTC" else "xau_csv")
    files  = sorted(glob.glob(os.path.join(folder, f"{sym}USDT-{tf}-2026-*.csv")))
    frames = []
    for f in files:
        with open(f) as fh: first = fh.readline()
        hdr = first.startswith("open_time")
        d = pd.read_csv(f, header=0 if hdr else None, names=None if hdr else COLS)
        frames.append(d)
    df = pd.concat(frames, ignore_index=True).iloc[:, :6]
    df.columns = ["open_time","open","high","low","close","volume"]
    ts = df["open_time"].astype(np.int64)
    if ts.iloc[0] > 1e15: ts = ts // 1000
    df["timestamp"] = pd.to_datetime(ts, unit="ms", utc=True).dt.tz_localize(None)
    df = df.drop(columns=["open_time"]).set_index("timestamp").sort_index().astype(float)
    return df[~df.index.duplicated(keep="first")]

def _ema(s,n): return s.ewm(span=n,adjust=False).mean()
def _sma(s,n): return s.rolling(n).mean()
def _wma(s,n):
    w=np.arange(1,n+1,dtype=float)
    return s.rolling(n).apply(lambda x:np.dot(x,w)/w.sum(),raw=True)
def _hma(s,n):
    sqrt_n=max(2,int(round(n**0.5)))
    return _wma(2*_wma(s,n//2)-_wma(s,n),sqrt_n)
def _rsi(s,n=14):
    d=s.diff(); g=d.clip(lower=0).ewm(alpha=1/n,adjust=False).mean()
    l=(-d.clip(upper=0)).ewm(alpha=1/n,adjust=False).mean()
    rs=g/l.replace(0,np.nan); r=100-100/(1+rs)
    r=r.where(~((l==0)&(g>0)),100.0); return r.fillna(50.0)
def _rma(s,n): return s.ewm(alpha=1/n,adjust=False).mean()
def _tr(df):
    pc=df["close"].shift(1)
    return pd.concat([df["high"]-df["low"],(df["high"]-pc).abs(),(df["low"]-pc).abs()],axis=1).max(axis=1)
def _atr(df,n=14): return _rma(_tr(df),n)
def _adx(df,n=14):
    up=df["high"].diff(); dn=-df["low"].diff()
    pdm=np.where((up>dn)&(up>0),up,0.); mdm=np.where((dn>up)&(dn>0),dn,0.)
    tr_r=_rma(_tr(df),n)
    pdi=100*_rma(pd.Series(pdm,index=df.index),n)/tr_r
    mdi=100*_rma(pd.Series(mdm,index=df.index),n)/tr_r
    dx=100*(pdi-mdi).abs()/(pdi+mdi).replace(0,np.nan)
    return _rma(dx.fillna(0),n)
def _mtf_bias(c15,c1h,c4h):
    def vote(s):
        ef,es,r=_ema(s,20),_ema(s,50),_rsi(s,14)
        v=(np.where(s>ef,1,-1)+np.where(ef>es,1,-1)+np.where(r>55,1,np.where(r<45,-1,0)))
        return pd.Series(v,index=s.index)
    return (sum([vote(c15),vote(c1h),vote(c4h)])/3)/3.0*100.0
def _merge(df_b,dfh,freq,pfx):
    h=dfh.copy(); h["avail_at"]=(h.index+pd.Timedelta(freq)).astype("datetime64[ns]")
    left=df_b.reset_index()[["timestamp"]].copy()
    left["timestamp"]=left["timestamp"].astype("datetime64[ns]")
    right=h.reset_index().rename(columns={"timestamp":f"{pfx}_ot"})
    m=pd.merge_asof(left,right,left_on="timestamp",right_on="avail_at",direction="backward")
    return m.set_index("timestamp")[["open","high","low","close","volume"]].add_prefix(f"{pfx}_")


def compute_signals(df15, df1h, df4h, cfg):
    out = df15.copy()
    out = out.join(_merge(df15,df1h,"1h","h1")).join(_merge(df15,df4h,"4h","h4"))
    ef4,es4 = _ema(out["h4_close"],20), _ema(out["h4_close"],50)
    ef1,es1 = _ema(out["h1_close"],20), _ema(out["h1_close"],50)

    bias = _mtf_bias(out["close"],out["h1_close"],out["h4_close"]); out["bias"]=bias
    bg = cfg["bias_gate"]
    bias_l = bias > bg; bias_s = bias < -bg

    pull_pct = cfg["pullback_pct"]
    near  = (out["h1_close"]-ef1).abs()/ef1 <= pull_pct
    wick_l = (out["h1_low"]  <= ef1*1.003) & (out["h1_close"]>ef1)
    wick_s = (out["h1_high"] >= ef1*0.997) & (out["h1_close"]<ef1)
    pull_l = near|wick_l; pull_s = near|wick_s

    adx = _adx(out,14); out["adx"]=adx
    atr = _atr(out,14); out["atr"]=atr
    rising = adx>adx.shift(1).fillna(0)
    adx_ok = (adx>cfg["adx_min"])&(adx<=cfg["adx_max"])&rising

    if cfg["score_mode"] == "classic":
        e9=_ema(out["close"],9); e20=_ema(out["close"],20)
        rsi15=_rsi(out["close"],14)
        volma=_sma(out["volume"],20); vol_ok=(volma>0)&(out["volume"]>=volma)
        sc_l=((out["close"]>e9).astype(float)+(out["close"]>e20).astype(float)
              +((rsi15>=35)&(rsi15<=75)).astype(float)+vol_ok.astype(float))
        sc_s=((out["close"]<e9).astype(float)+(out["close"]<e20).astype(float)
              +((rsi15>=28)&(rsi15<=65)).astype(float)+vol_ok.astype(float))
    else:  # sj hybrid
        ema5=_ema(out["close"],5); sma9=_sma(out["close"],9)
        hma20=_hma(out["close"],20)
        hh10=out["high"].rolling(10).max().shift(1); ll10=out["low"].rolling(10).min().shift(1)
        rsi15=_rsi(out["close"],14)
        volma=_sma(out["volume"],20); vol_ok=(volma>0)&(out["volume"]>=volma)
        sc_l=((out["close"]>hma20).astype(float)+(ema5>sma9).astype(float)
             +((rsi15>=35)&(rsi15<=75)).astype(float)+vol_ok.astype(float)
             +(out["close"]>hh10).astype(float))
        sc_s=((out["close"]<hma20).astype(float)+(ema5<sma9).astype(float)
             +((rsi15>=28)&(rsi15<=65)).astype(float)+vol_ok.astype(float)
             +(out["close"]<ll10).astype(float))
    out["score_l"]=sc_l; out["score_s"]=sc_s
    ms=cfg["min_score"]

    raw_b=(bias_l&pull_l&adx_ok&(sc_l>=ms)).fillna(False)
    raw_s=(bias_s&pull_s&adx_ok&(sc_s>=ms)).fillna(False)

    cd=cfg["cooldown"]
    sb=pd.Series(False,index=out.index); ss=pd.Series(False,index=out.index)
    blk=0
    for i in range(len(out)):
        if blk>0: blk-=1; continue
        if raw_b.iloc[i]: sb.iloc[i]=True; blk=cd
        elif raw_s.iloc[i]: ss.iloc[i]=True; blk=cd
    out["buy"]=sb; out["sell"]=ss
    return out


def backtest_detailed(df, cfg, notional=1000, sl_mult=1.2, fee=0.0004, warmup=500):
    """Returns (trades_df, summary_dict). trades_df is one row per closed entry
    (aggregating TP1 partial + final exit into a single trade record)."""
    tp1_r=cfg["tp1_r"]; tp1_frac=cfg["tp1_frac"]; tp2_r=cfg["tp2_r"]
    rows=[]; pos=None; arr=df.reset_index()
    for i in range(warmup,len(arr)-1):
        row=arr.iloc[i]; nxt=arr.iloc[i+1]
        if pos is None:
            sig="long" if row["buy"] else ("short" if row["sell"] else None)
            if sig:
                entry=float(nxt["open"]); dist=float(row["atr"])*sl_mult
                dist=max(dist,entry*0.012); dist=min(dist,entry*0.035)
                sl=entry-dist if sig=="long" else entry+dist
                tp1=entry+tp1_r*dist if sig=="long" else entry-tp1_r*dist
                tp2=entry+tp2_r*dist if sig=="long" else entry-tp2_r*dist
                pos=dict(side=sig,entry=entry,sl=sl,tp1=tp1,tp2=tp2,dist=dist,
                         tp1_hit=False,entry_time=nxt["timestamp"],
                         entry_score=float(row["score_l"] if sig=="long" else row["score_s"]),
                         entry_adx=float(row["adx"]),entry_bias=float(row["bias"]),
                         tp1_pnl=0.0)
        else:
            hi=float(nxt["high"]); lo=float(nxt["low"]); s=pos["side"]
            if not pos["tp1_hit"]:
                if (s=="long" and hi>=pos["tp1"]) or (s=="short" and lo<=pos["tp1"]):
                    p1=((pos["tp1"]-pos["entry"])/pos["entry"] if s=="long"
                        else (pos["entry"]-pos["tp1"])/pos["entry"])
                    pos["tp1_pnl"]=p1*notional*tp1_frac-fee*notional
                    pos["tp1_hit"]=True; pos["sl"]=pos["entry"]
            sl_hit=(s=="long" and lo<=pos["sl"]) or (s=="short" and hi>=pos["sl"])
            tp2_hit=(s=="long" and hi>=pos["tp2"]) or (s=="short" and lo<=pos["tp2"])
            exit_now=None; reason=None; exit_px=None
            if sl_hit:
                exit_px=pos["sl"]
                reason="BE" if (pos["tp1_hit"] and abs(exit_px-pos["entry"])<1e-9) else "SL"
                exit_now=True
            elif tp2_hit:
                exit_px=pos["tp2"]; reason="TP2"; exit_now=True
            if exit_now:
                pnl=((exit_px-pos["entry"])/pos["entry"] if s=="long"
                     else (pos["entry"]-exit_px)/pos["entry"])
                frac=(1.0-tp1_frac) if pos["tp1_hit"] else 1.0
                runner_pnl=pnl*notional*frac-fee*notional
                total_pnl=pos["tp1_pnl"]+runner_pnl
                r_mult=((exit_px-pos["entry"])/pos["dist"] if s=="long"
                        else (pos["entry"]-exit_px)/pos["dist"])
                rows.append({
                    "EntryTime":pos["entry_time"],"ExitTime":nxt["timestamp"],
                    "Side":s.upper(),"Entry":round(pos["entry"],2),
                    "Exit":round(exit_px,2),"SL":round(pos["entry"]-pos["dist"] if s=="long"
                                                       else pos["entry"]+pos["dist"],2),
                    "TP1":round(pos["tp1"],2),"TP2":round(pos["tp2"],2),
                    "TP1_Hit":"Yes" if pos["tp1_hit"] else "No",
                    "ExitReason":reason,"R_Multiple":round(r_mult,2),
                    "Score":round(pos["entry_score"],0),"ADX":round(pos["entry_adx"],1),
                    "Bias":round(pos["entry_bias"],0),
                    "PnL_USD":round(total_pnl,2),
                    "BarsHeld":int((nxt["timestamp"]-pos["entry_time"]).total_seconds()/900),
                })
                pos=None
    tdf=pd.DataFrame(rows)
    if len(tdf)==0:
        return tdf, dict(trades=0,wr=0,pnl=0,maxdd=0,avg_win=0,avg_loss=0,pf=0,
                         tp2_pct=0,be_pct=0,sl_pct=0,tr_month=0)
    pnls=tdf["PnL_USD"].values
    cum=np.cumsum(pnls); tdf["CumPnL"]=np.round(cum,2)
    dd=float(np.min(cum-np.maximum.accumulate(cum)))
    n=len(tdf); wins=tdf[tdf["PnL_USD"]>0]; losses=tdf[tdf["PnL_USD"]<=0]
    gross_win=wins["PnL_USD"].sum(); gross_loss=abs(losses["PnL_USD"].sum())
    summary=dict(
        trades=n, wr=round(len(wins)/n*100,1), pnl=round(pnls.sum(),2),
        maxdd=round(dd,2),
        avg_win=round(wins["PnL_USD"].mean(),2) if len(wins) else 0,
        avg_loss=round(losses["PnL_USD"].mean(),2) if len(losses) else 0,
        pf=round(gross_win/gross_loss,2) if gross_loss>0 else 0,
        tp2_pct=round((tdf["ExitReason"]=="TP2").sum()/n*100,1),
        be_pct=round((tdf["ExitReason"]=="BE").sum()/n*100,1),
        sl_pct=round((tdf["ExitReason"]=="SL").sum()/n*100,1),
        tr_month=round(n/5,1),
    )
    return tdf, summary


# ── Configs (matching LIVE Railway settings) ──────────────────────────────────
NOTIONAL = 50 * 20
CFG_A = dict(name="A Current FAST v2 (LIVE)", score_mode="classic",
             bias_gate=60, pullback_pct=0.025, adx_min=18, adx_max=50,
             min_score=4, cooldown=4, tp1_r=0.5, tp1_frac=0.40, tp2_r=3.0)
CFG_L = dict(name="L SJ Hybrid (PROPOSED)", score_mode="sj",
             bias_gate=60, pullback_pct=0.025, adx_min=18, adx_max=50,
             min_score=4, cooldown=4, tp1_r=0.5, tp1_frac=0.40, tp2_r=3.0)

results = {}
trades_all = {}
for sym in ["BTC","XAU"]:
    df15=load_tf(sym,"15m"); df1h=load_tf(sym,"1h"); df4h=load_tf(sym,"4h")
    for cfg,tag in [(CFG_A,"A"),(CFG_L,"L")]:
        d=compute_signals(df15,df1h,df4h,cfg)
        tdf,summ=backtest_detailed(d,cfg,notional=NOTIONAL)
        results[(tag,sym)]=summ
        trades_all[(tag,sym)]=tdf
        print(f"  {cfg['name']:<28} {sym}: {summ['trades']} trades, "
              f"WR {summ['wr']}%, PnL ${summ['pnl']}, PF {summ['pf']}")

# ── Build Summary sheet ────────────────────────────────────────────────────────
summary_rows=[]
for tag,label in [("A","A Current FAST v2 (LIVE NOW)"),("L","L SJ Hybrid (PROPOSED)")]:
    for sym in ["BTC","XAU"]:
        s=results[(tag,sym)]
        summary_rows.append({
            "Config":label,"Symbol":sym,"Trades":s["trades"],
            "Trades/Month":s["tr_month"],"WinRate%":s["wr"],
            "NetPnL_USD":s["pnl"],"ProfitFactor":s["pf"],
            "MaxDD_USD":s["maxdd"],"AvgWin":s["avg_win"],"AvgLoss":s["avg_loss"],
            "TP2%":s["tp2_pct"],"BE%":s["be_pct"],"SL%":s["sl_pct"],
        })
    # combined
    sb=results[(tag,"BTC")]; sx=results[(tag,"XAU")]
    tot=sb["trades"]+sx["trades"]
    wr=(sb["wr"]*sb["trades"]+sx["wr"]*sx["trades"])/tot if tot else 0
    summary_rows.append({
        "Config":label,"Symbol":"COMBINED","Trades":tot,
        "Trades/Month":round(tot/5,1),"WinRate%":round(wr,1),
        "NetPnL_USD":round(sb["pnl"]+sx["pnl"],2),"ProfitFactor":"",
        "MaxDD_USD":min(sb["maxdd"],sx["maxdd"]),"AvgWin":"","AvgLoss":"",
        "TP2%":"","BE%":"","SL%":"",
    })
summary_df=pd.DataFrame(summary_rows)

# ── Monthly breakdown ──────────────────────────────────────────────────────────
monthly_rows=[]
for tag,label in [("A","Current FAST v2"),("L","SJ Hybrid")]:
    for sym in ["BTC","XAU"]:
        tdf=trades_all[(tag,sym)]
        if len(tdf)==0: continue
        tdf2=tdf.copy(); tdf2["Month"]=pd.to_datetime(tdf2["EntryTime"]).dt.strftime("%Y-%m")
        for month,g in tdf2.groupby("Month"):
            monthly_rows.append({
                "Config":label,"Symbol":sym,"Month":month,
                "Trades":len(g),"WinRate%":round((g["PnL_USD"]>0).sum()/len(g)*100,1),
                "NetPnL_USD":round(g["PnL_USD"].sum(),2),
            })
monthly_df=pd.DataFrame(monthly_rows)

# ── Write XLSX ─────────────────────────────────────────────────────────────────
with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as xw:
    summary_df.to_excel(xw, sheet_name="Summary", index=False)
    monthly_df.to_excel(xw, sheet_name="Monthly", index=False)
    for tag in ["A","L"]:
        for sym in ["BTC","XAU"]:
            tdf=trades_all[(tag,sym)]
            sheet=f"Trades_{tag}_{sym}"
            if len(tdf)==0:
                pd.DataFrame({"note":["no trades"]}).to_excel(xw,sheet_name=sheet,index=False)
            else:
                tdf.to_excel(xw,sheet_name=sheet,index=False)

# ── Auto-fit column widths + header styling ────────────────────────────────────
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
wb=load_workbook(OUT_PATH)
hdr_fill=PatternFill(start_color="1F4E78",end_color="1F4E78",fill_type="solid")
hdr_font=Font(color="FFFFFF",bold=True)
for ws in wb.worksheets:
    for cell in ws[1]:
        cell.fill=hdr_fill; cell.font=hdr_font
        cell.alignment=Alignment(horizontal="center")
    for col in ws.columns:
        width=max((len(str(c.value)) for c in col if c.value is not None),default=10)
        ws.column_dimensions[col[0].column_letter].width=min(width+2,22)
    ws.freeze_panes="A2"
wb.save(OUT_PATH)

print(f"\n✅ Saved: {OUT_PATH}")
print(f"   Sheets: Summary, Monthly, Trades_A_BTC, Trades_A_XAU, Trades_L_BTC, Trades_L_XAU")
print(f"   Total trades logged: A={len(trades_all[('A','BTC')])+len(trades_all[('A','XAU')])}, "
      f"L={len(trades_all[('L','BTC')])+len(trades_all[('L','XAU')])}")
